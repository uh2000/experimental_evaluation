"""
Deterministic Equivalent (DE) for Salmon Farming MILP — extensive form.

One copy of ip.py's exact MILP formulation per scenario, tied together by
non-anticipativity constraints (NAC).  The formulation matches ip.py exactly:

  stk      BINARY
  q        SEMICONT [1, per_unit_qmax] — either 0 (not stocked) or in [1, UB] (stocked)
  harv     BINARY
  x        CONTINUOUS = harv × q via McCormick (convex hull, no big-M slack)
  a_new    CONTINUOUS [0,1], propagated by equality — always integral
  e_alive  CONTINUOUS [0,1], incremental propagation
  h_exist  BINARY
  H_exist_biom  CONTINUOUS — keeps large biomass_exist out of the objective

  Pricing: weight-class table (DEFAULT_PRICES in ip.py) — 9 brackets from 1-2 kg
           up to 9+ kg, peaking at the 5-6 kg class.

  Biomass in density/MAB:  bpf * (q − Σ x)  directly (no B_cohort auxiliary)
  Location fallow:         a_new == 0 / e_alive == 0  (same as ip.py)
  Objective:               bpf[s][t] * x * price  (via x = harv*q McCormick)

Scenario tree (2 stages of uncertainty, T = 60 months):
  Stage 0 (months  0-29): 1  node                  — all 9 scenarios share one plan
  Stage 1 (months 30-59): 3  nodes  (t1)            —  3 scenarios each
Total: 3^2 = 9 scenarios.

NAC enforced on: stk, q, harv (harvest months in stage), h_exist (in stage).
Derived variables (a_new, x) satisfy NAC automatically via equality propagation.
Stage-0 has 1 node: t1 is NOT known when months 0–29 decisions are made.
"""

from collections import defaultdict
from time import time

import matplotlib
import matplotlib.pyplot as plt
from matplotlib.lines import Line2D as _L2D
from matplotlib.patches import Patch as _MPatch
import numpy as np
import pandas as pd
import gurobipy as gb
from gurobipy import GRB, quicksum

from IP import SalmonFarmingMILP
from instance import units_df, loc_mab, regional_mab


class DeterministicEquivalent:
    """Deterministic equivalent (extensive form) for 9-scenario 2-stage salmon farming."""

    def __init__(
        self,
        units_df: pd.DataFrame,
        loc_mab: dict,
        regional_mab: float,
        *,
        T: int = 60,
        mip_gap: float = 0.02,
        time_limit: float = None,
        seed: int = None,
        temps_normal: np.ndarray = None,
        temps_bad: np.ndarray = None,
        temps_good: np.ndarray = None,
    ):
        self.units_df     = units_df
        self.loc_mab      = loc_mab
        self.regional_mab = regional_mab
        self.T            = T
        self.mip_gap      = mip_gap
        self.time_limit   = time_limit
        self.seed         = seed

        # Stage boundaries — 2 evenly spaced 30-month stages
        self.stage0_months = list(range(0, 30))
        self.stage1_months = list(range(30, T))

        days = 30
        self.temps_normal = (temps_normal if temps_normal is not None
                             else np.array([5, 5, 5, 6, 9, 12, 14, 16.5, 15.5, 13, 10, 7.5]))
        self.temps_bad    = (temps_bad    if temps_bad    is not None
                             else np.array([3, 3, 3, 4,  7, 10, 12, 14.5, 13.5, 11,  8, 5.5]))
        self.temps_good   = (temps_good   if temps_good   is not None
                             else np.array([7, 7, 7, 8, 11, 14, 16, 18.5, 17.5, 15, 12, 9.5]))

        self.S_normal = (1.0 - 0.0002) ** days
        self.S_bad    = (1.0 - 0.001)  ** days

        self.stage_labels = ["bad", "normal", "good"]

        self.milp_objects:    list = []
        self.scenario_names:  list = []
        self.scenario_bundle: list = []   # stage-0 node: t1 index (0/1/2)
        self.probs:           np.ndarray = None
        self.n_scenarios:     int = 0
        self.model:           gb.Model = None

        self._stk          = None
        self._q            = None
        self._harv         = None
        self._x            = None
        self._a_new        = None
        self._e_alive      = None
        self._h_exist      = None
        self._H_exist_biom = None

        self.obj_val    = None
        self.solve_time = None
        self.status     = None

    @staticmethod
    def _tile(arr12: np.ndarray, n: int) -> np.ndarray:
        return np.tile(arr12, (n // 12) + 1)[:n]

    def _make_scenario_data(self):
        """Build one SalmonFarmingMILP per scenario."""
        temp_map = {
            "bad":    self.temps_bad,
            "normal": self.temps_normal,
            "good":   self.temps_good,
        }
        for l1 in self.stage_labels:
            for l2 in self.stage_labels:
                name = f"s1_{l1}__s2_{l2}"

                temps_t = np.zeros(self.T)
                S_t     = np.full(self.T, self.S_normal)

                stage_seq = [
                    (l1, self.stage0_months),
                    (l2, self.stage1_months),
                ]
                for i, (sl, months) in enumerate(stage_seq):
                    prev = stage_seq[i - 1][0] if i > 0 else None
                    surv = (self.S_bad if (sl == "good" and prev == "good")
                            else self.S_normal)
                    S_t[months] = surv
                    temps_t[months] = self._tile(temp_map[sl], len(months))

                milp = SalmonFarmingMILP(
                    units_df=self.units_df,
                    temps_t=temps_t,
                    survival_rates=S_t,
                    loc_mab=self.loc_mab,
                    regional_mab=self.regional_mab,
                    horizon_months=self.T,
                    scenario_name=name,
                )
                self.milp_objects.append(milp)
                self.scenario_names.append(name)
                self.scenario_bundle.append(self.stage_labels.index(l1))

        self.n_scenarios = len(self.milp_objects)
        self.probs = np.full(self.n_scenarios, 1.0 / self.n_scenarios)
        print(f"  Prepared {self.n_scenarios} scenarios  (3×3, 2 stages of uncertainty)")

    def build(self):
        """Assemble the full DE model on the extensive form, matching ip.py exactly."""
        t_build = time()
        self._make_scenario_data()

        ref          = self.milp_objects[0]
        U            = ref.U
        U_exist      = ref.U_exist
        U_exist_set  = set(U_exist)
        Tset         = ref.Tset
        A_by_ut      = ref.A_by_ut
        A_harv_set   = ref.A_set

        earliest_ht_union = {
            s: min(milp.earliest_harvest_t[s] for milp in self.milp_objects)
            for s in Tset
        }
        H_union = [
            (u, s, t)
            for u in U
            for s in Tset
            for t in Tset
            if earliest_ht_union[s] <= t <= s + ref.maturation_t
        ]
        H_set   = set(H_union)
        H_by_us = defaultdict(list)
        H_by_ut = defaultdict(list)
        for u, s, t in H_union:
            H_by_us[(u, s)].append(t)
            H_by_ut[(u, t)].append(s)
        vol          = ref.vol
        loc_of       = ref.loc_of
        L_list       = ref.L_list

        density_limit          = ref.density_limit
        smolt_cost             = ref.smolt_cost_per_head
        terminal_value_per_kg  = ref.terminal_value_per_kg
        feed_cost_per_kg_month = ref.feed_cost_per_kg_month
        df_disc                = ref.df
        L_MIN_EXIST            = ref.L_MIN_EXIST
        L_MAX_EXIST            = ref.L_MAX_EXIST
        allowed_et             = set(range(L_MIN_EXIST, L_MAX_EXIST + 1))
        cap_kg                 = {u: density_limit * vol[u] for u in U}

        S     = range(self.n_scenarios)
        Q_MAX = 1_000_000
        T_end = self.T - 1
        FALLOW = 2

        for milp in self.milp_objects:
            if milp.model is not None:
                milp.model.dispose()
                milp.model = None

        prev_x_lt = defaultdict(list)
        for u, s, t in A_harv_set:
            for tau in H_by_us[(u, s)]:
                if tau < t:
                    prev_x_lt[(u, t)].append((s, tau))

        per_unit_qmax = {}
        for sc in S:
            milp = self.milp_objects[sc]
            for u in U:
                for s in Tset:
                    harvest_ts = H_by_us.get((u, s), [])
                    if not harvest_ts:
                        per_unit_qmax[(sc, u, s)] = 0
                    else:
                        horizon_end = min(s + ref.maturation_t, T_end)
                        max_sbpf = max(
                            (float(milp.sbpf[s][t]) for t in range(s, horizon_end + 1)),
                            default=1e-9,
                        )
                        max_sbpf = max(max_sbpf, 1e-9)
                        per_unit_qmax[(sc, u, s)] = max(int(np.floor(cap_kg[u] / max_sbpf)), 0)

        m = gb.Model("DE_Salmon_9sc")
        m.Params.OutputFlag = 1

        stk = m.addVars(S, U, Tset, vtype=GRB.BINARY, name="z")
        q   = m.addVars(S, U, Tset, lb=1.0, ub=Q_MAX, vtype=GRB.SEMICONT, name="q")

        harv_keys = [(sc, u, s, t) for sc in S for (u, s, t) in H_set]
        harv = m.addVars(harv_keys, vtype=GRB.BINARY, name="harv")
        x    = m.addVars(harv_keys, lb=0.0, name="x")

        anew_keys = [(sc, u, s, t) for sc in S for (u, s, t) in A_harv_set]
        a_new = m.addVars(anew_keys, lb=0.0, ub=1.0, name="a_new")

        e_alive    = m.addVars(S, U_exist, Tset, lb=0.0, ub=1.0, name="e_alive")
        h_exist    = m.addVars(S, U_exist, Tset, vtype=GRB.BINARY, name="h_exist")
        H_exist_biom = m.addVars(
            [(sc, u, t) for sc in S for u in U_exist for t in sorted(allowed_et)],
            lb=0.0, name="H_exist_biom",
        )

        BIOMASS_PENALTY = 1_000
        slack_density = m.addVars(S, U,      Tset, lb=0.0, name="slack_density")
        slack_mab_loc = m.addVars(S, L_list, Tset, lb=0.0, name="slack_mab_loc")
        slack_mab_reg = m.addVars(S, Tset,        lb=0.0, name="slack_mab_reg")

        print(f"  Variables: {m.NumVars:,}")

        for sc in S:
            for u in U:
                for s in Tset:
                    qmax = per_unit_qmax[(sc, u, s)]
                    q[sc, u, s].UB = qmax
                    if qmax == 0:
                        stk[sc, u, s].UB = 0
                    else:
                        m.addConstr(q[sc, u, s] <= qmax * stk[sc, u, s])

        for sc in S:
            milp = self.milp_objects[sc]

            for u in U:
                for s in Tset:
                    allowed_ts = H_by_us[(u, s)]
                    if not allowed_ts:
                        m.addConstr(stk[sc, u, s] == 0)
                    else:
                        m.addConstr(
                            quicksum(harv[sc, u, s, t] for t in allowed_ts) <= stk[sc, u, s]
                        )

            for u in U:
                for s in Tset:
                    sc_eht = milp.earliest_harvest_t[s]
                    for t in H_by_us[(u, s)]:
                        if t < sc_eht:
                            m.addConstr(harv[sc, u, s, t] == 0)

            for u, s, t in H_set:
                qmax = per_unit_qmax[(sc, u, s)]
                m.addConstr(x[sc, u, s, t] <= q[sc, u, s])
                m.addConstr(x[sc, u, s, t] <= qmax * harv[sc, u, s, t])
                m.addConstr(x[sc, u, s, t] >= q[sc, u, s] - qmax * (1 - harv[sc, u, s, t]))

            for u in U:
                for s in Tset:
                    if (u, s, s) in A_harv_set:
                        m.addConstr(a_new[sc, u, s, s] == stk[sc, u, s])
            for u, s, t in A_harv_set:
                if t > s:
                    if (u, s, t - 1) in H_set:
                        m.addConstr(
                            a_new[sc, u, s, t] == a_new[sc, u, s, t - 1]
                            - harv[sc, u, s, t - 1]
                        )
                    else:
                        m.addConstr(a_new[sc, u, s, t] == a_new[sc, u, s, t - 1])

            for u in U_exist:
                m.addConstr(quicksum(h_exist[sc, u, t] for t in allowed_et) <= 1)
                for t in Tset:
                    if t not in allowed_et:
                        m.addConstr(h_exist[sc, u, t] == 0)

            for u in U_exist:
                for t in allowed_et:
                    b_t = float(milp.biomass_exist[u][t])
                    H_exist_biom[sc, u, t].UB = b_t
                    m.addConstr(H_exist_biom[sc, u, t] == b_t * h_exist[sc, u, t])

            for u in U_exist:
                m.addConstr(e_alive[sc, u, 0] == 1)
                for t in range(1, self.T):
                    m.addConstr(
                        e_alive[sc, u, t] == e_alive[sc, u, t - 1] - h_exist[sc, u, t - 1]
                    )

            for u in U:
                for t in Tset:
                    alive_harv = quicksum(a_new[sc, u, s, t] for s in A_by_ut[(u, t)])
                    if u in U_exist_set:
                        m.addConstr(alive_harv + e_alive[sc, u, t] <= 1)
                    else:
                        m.addConstr(alive_harv <= 1)

            for u in U:
                for t in Tset:
                    harvest_terms = []
                    for offset in range(1, FALLOW + 1):
                        tau = t - offset
                        if tau < 0:
                            continue
                        harvest_terms.extend(harv[sc, u, s, tau] for s in H_by_ut[(u, tau)])
                        if u in U_exist_set and tau in allowed_et:
                            harvest_terms.append(h_exist[sc, u, tau])
                    if harvest_terms:
                        m.addConstr(stk[sc, u, t] + quicksum(harvest_terms) <= 1)

            loc_fallow_start = {
                "Location 1": 12, "Location 2": 13, "Location 3": 14,
                "Location 4": 15, "Location 5": 16, "Location 6": 18,
            }
            LOC_FALLOW_DURATION = 2
            LOC_FALLOW_CYCLE    = 24
            for ell, fallow_t0 in loc_fallow_start.items():
                units_in_loc = [u for u in U if loc_of[u] == ell]
                fallow_months = []
                ft = fallow_t0
                while ft < self.T:
                    for offset in range(LOC_FALLOW_DURATION):
                        if ft + offset < self.T:
                            fallow_months.append(ft + offset)
                    ft += LOC_FALLOW_CYCLE
                for u in units_in_loc:
                    for t in fallow_months:
                        if u in U_exist_set:
                            m.addConstr(e_alive[sc, u, t] == 0)
                        for s in A_by_ut[(u, t)]:
                            m.addConstr(a_new[sc, u, s, t] == 0)

            for u in U:
                for t in Tset:
                    biom_new = (
                        quicksum(float(milp.sbpf[s][t]) * q[sc, u, s] for s in A_by_ut[(u, t)])
                        - quicksum(float(milp.sbpf[s][t]) * x[sc, u, s, tau]
                                   for s, tau in prev_x_lt[(u, t)])
                    )
                    biom_old = (
                        float(milp.biomass_exist[u][t]) * e_alive[sc, u, t]
                        if u in U_exist_set else 0.0
                    )
                    m.addConstr(biom_new + biom_old <= cap_kg[u] + slack_density[sc, u, t])

            for ell in L_list:
                units_ell = [u for u in U if loc_of[u] == ell]
                for t in Tset:
                    biom_loc = quicksum(
                        quicksum(float(milp.sbpf[s][t]) * q[sc, u, s] for s in A_by_ut[(u, t)])
                        - quicksum(float(milp.sbpf[s][t]) * x[sc, u, s, tau]
                                   for s, tau in prev_x_lt[(u, t)])
                        + (float(milp.biomass_exist[u][t]) * e_alive[sc, u, t]
                           if u in U_exist_set else 0.0)
                        for u in units_ell
                    )
                    m.addConstr(biom_loc <= self.loc_mab[ell] + slack_mab_loc[sc, ell, t])

            for t in Tset:
                biom_reg = quicksum(
                    quicksum(float(milp.sbpf[s][t]) * q[sc, u, s] for s in A_by_ut[(u, t)])
                    - quicksum(float(milp.sbpf[s][t]) * x[sc, u, s, tau]
                               for s, tau in prev_x_lt[(u, t)])
                    + (float(milp.biomass_exist[u][t]) * e_alive[sc, u, t]
                       if u in U_exist_set else 0.0)
                    for u in U
                )
                m.addConstr(biom_reg <= self.regional_mab + slack_mab_reg[sc, t])

        print(f"  Per-scenario constraints: {m.NumConstrs:,}")

        # ── Non-anticipativity constraints ────────────────────────────────────
        stage0_set = set(self.stage0_months)
        stage1_set = set(self.stage1_months)

        def _add_nac(month_set, node_key_fn):
            nodes = {}
            for sc_idx in S:
                key = node_key_fn(sc_idx)
                nodes.setdefault(key, []).append(sc_idx)
            for sc_list in nodes.values():
                rep = sc_list[0]
                for sc in sc_list[1:]:
                    for u in U:
                        for t in month_set:
                            m.addConstr(stk[sc, u, t] == stk[rep, u, t])
                            m.addConstr(q[sc, u, t]   == q[rep, u, t])
                    for u in U:
                        for s in Tset:
                            for th in H_by_us.get((u, s), []):
                                if th in month_set:
                                    m.addConstr(harv[sc, u, s, th] == harv[rep, u, s, th])
                    for u in U_exist:
                        for t in month_set:
                            m.addConstr(h_exist[sc, u, t] == h_exist[rep, u, t])

        # Stage 0 — 1 node: all 9 scenarios share one plan
        _add_nac(stage0_set, lambda _: 0)
        # Stage 1 — 3 nodes: scenarios sharing t1 agree on months 30–59
        _add_nac(stage1_set, lambda i: self.scenario_bundle[i])

        print(f"  Total constraints with NAC: {m.NumConstrs:,}")

        terminal_coeff = terminal_value_per_kg * df_disc[T_end]

        obj = gb.LinExpr()
        for sc in S:
            milp = self.milp_objects[sc]
            p    = float(self.probs[sc])

            def _price_new(s, t, _milp=milp):
                return _milp._price_for_kg(_milp.wpf[s][t])

            def _price_exist(u, t, _milp=milp):
                return _milp._price_for_kg(_milp.wpf_exist[u][t])

            for u, s, t in H_set:
                obj.add(x[sc, u, s, t],
                        p * float(milp.sbpf[s][t]) * _price_new(s, t) * df_disc[t])

            for u in U_exist:
                for t in sorted(allowed_et):
                    obj.add(H_exist_biom[sc, u, t],
                            p * _price_exist(u, t) * df_disc[t])

            for u in U:
                for s in Tset:
                    obj.add(q[sc, u, s], -p * smolt_cost * df_disc[s])

            for u in U:
                for t in Tset:
                    fc_dt = p * feed_cost_per_kg_month * df_disc[t]
                    for s in A_by_ut[(u, t)]:
                        obj.add(q[sc, u, s], -fc_dt * float(milp.sbpf[s][t]))
                    for s, tau in prev_x_lt[(u, t)]:
                        obj.add(x[sc, u, s, tau], fc_dt * float(milp.sbpf[s][t]))
                    if u in U_exist_set:
                        obj.add(e_alive[sc, u, t],
                                -fc_dt * float(milp.biomass_exist[u][t]))

            for u in U:
                for s in A_by_ut[(u, T_end)]:
                    bpf_end = float(milp.sbpf[s][T_end])
                    obj.add(q[sc, u, s], p * bpf_end * terminal_coeff)
                    for tau in H_by_us[(u, s)]:
                        if tau <= T_end:
                            obj.add(x[sc, u, s, tau], -p * bpf_end * terminal_coeff)

            for u in U_exist:
                b_end = float(milp.biomass_exist[u][T_end])
                obj.addConstant(p * b_end * terminal_coeff)
                for tau in sorted(allowed_et):
                    b_tau = float(milp.biomass_exist[u][tau])
                    ratio = b_end / b_tau if b_tau > 1e-9 else 0.0
                    obj.add(H_exist_biom[sc, u, tau], -p * ratio * terminal_coeff)

        for sc in S:
            p_sc = float(self.probs[sc])
            for u in U:
                for t in Tset:
                    obj.add(slack_density[sc, u, t], -p_sc * BIOMASS_PENALTY)
            for ell in L_list:
                for t in Tset:
                    obj.add(slack_mab_loc[sc, ell, t], -p_sc * BIOMASS_PENALTY)
            for t in Tset:
                obj.add(slack_mab_reg[sc, t], -p_sc * BIOMASS_PENALTY)

        m.setObjective(obj, GRB.MAXIMIZE)

        if self.mip_gap is not None:
            m.Params.MIPGap = self.mip_gap
        if self.time_limit is not None:
            m.Params.TimeLimit = self.time_limit
        if self.seed is not None:
            m.Params.Seed = self.seed

        self.model         = m
        self._stk          = stk
        self._q            = q
        self._harv         = harv
        self._x            = x
        self._a_new        = a_new
        self._e_alive      = e_alive
        self._h_exist      = h_exist
        self._H_exist_biom = H_exist_biom

        print(
            f"\nDE model built in {time() - t_build:.1f}s  "
            f"| {m.NumVars:,} vars  | {m.NumConstrs:,} constraints"
        )

    def solve(self):
        if self.model is None:
            raise RuntimeError("Call build() first.")
        print("\n--- Solving DE ---")
        t0 = time()
        self.model.optimize()
        self.solve_time = time() - t0
        self.status = self.model.Status
        if self.model.SolCount > 0:
            self.obj_val = self.model.ObjVal

    def print_results(self):
        ref     = self.milp_objects[0]
        U       = ref.U
        U_exist = ref.U_exist

        print("\n" + "=" * 70)
        status_map = {2: "OPTIMAL", 3: "INFEASIBLE", 5: "UNBOUNDED", 9: "TIME_LIMIT"}
        print(f"Status     : {status_map.get(self.status, str(self.status))}")
        print(f"Solve time : {self.solve_time:.1f}s")
        if self.model.SolCount > 0:
            print(f"Obj value  : {self.obj_val:,.2f}")
            print(f"MIP gap    : {self.model.MIPGap:.4%}")
        print("=" * 70)

        if self.model.SolCount == 0:
            return

        normal_sc = next(
            (i for i, nm in enumerate(self.scenario_names)
             if nm == "s1_normal__s2_normal"), 0)
        normal_ref = self.milp_objects[normal_sc]

        def _print_stage(label, months, sc):
            month_set = set(months)
            print(f"\n--- {label} ---")
            any_decision = False
            for u in U:
                for t in months:
                    if self._stk[sc, u, t].X > 0.5:
                        print(f"  stk[{u},{t}] = 1,  q = {self._q[sc, u, t].X:,.0f} fish")
                        any_decision = True
                for s_start in normal_ref.Tset:
                    for th in normal_ref.H_by_us.get((u, s_start), []):
                        if th in month_set:
                            if self._harv[sc, u, s_start, th].X > 0.5:
                                print(f"  harv[{u},stocked={s_start},harvest={th}] = 1"
                                      f"  q={self._q[sc, u, s_start].X:,.0f} fish")
                                any_decision = True
            for u in U_exist:
                for t in months:
                    if self._h_exist[sc, u, t].X > 0.5:
                        print(f"  h_exist[{u},{t}] = 1  (existing cohort harvested)")
                        any_decision = True
            if not any_decision:
                print("  (no decisions)")

        _print_stage("Stage-0 Decisions  (months 0-29, all-normal path)",
                     self.stage0_months, normal_sc)
        _print_stage("Stage-1 Decisions  (months 30-59, all-normal scenario: t1=t2=normal)",
                     self.stage1_months, normal_sc)

    def plot_biomass(self, output_file: str = "stats/de_biomass.png") -> "plt.Figure":
        """Plot expected biomass over time — one subplot per location + regional total."""
        matplotlib.use("Agg")

        if self.model is None or self.model.SolCount == 0:
            raise RuntimeError("No solution available. Call solve() before plot_biomass().")

        ref          = self.milp_objects[0]
        U            = ref.U
        U_exist      = ref.U_exist
        U_exist_set  = set(U_exist)
        Tset         = ref.Tset
        A_harv_set   = ref.A_set
        H_by_us      = ref.H_by_us
        loc_of       = ref.loc_of
        L_list       = ref.L_list
        months       = list(range(self.T))

        unit_biom = {}
        for sc in range(self.n_scenarios):
            milp = self.milp_objects[sc]
            unit_biom[sc] = {}
            for u in U:
                b = np.zeros(self.T)
                for s in Tset:
                    for t in Tset:
                        if (u, s, t) in A_harv_set:
                            harvested_through_t = sum(
                                self._x[sc, u, s, tau].X
                                for tau in H_by_us[(u, s)] if tau <= t
                                if (sc, u, s, tau) in self._x
                            )
                            b[t] += float(milp.sbpf[s][t]) * (
                                self._q[sc, u, s].X - harvested_through_t
                            )
                if u in U_exist_set:
                    b_ex = milp.biomass_exist
                    for t in Tset:
                        e_val = self._e_alive[sc, u, t].X
                        h_val = self._h_exist[sc, u, t].X
                        b[t] += b_ex[u][t] * (e_val - h_val)
                unit_biom[sc][u] = b / 1e3

        exp_biom = {
            u: sum(self.probs[sc] * unit_biom[sc][u] for sc in range(self.n_scenarios))
            for u in U
        }

        n_rows = len(L_list) + 1
        fig, axes = plt.subplots(n_rows, 1, figsize=(13, 3.2 * n_rows), sharex=True)
        if n_rows == 1:
            axes = [axes]

        cmap = plt.get_cmap("tab10")

        for i, loc in enumerate(L_list):
            ax = axes[i]
            units_in_loc = [u for u in U if loc_of[u] == loc]
            stack  = np.vstack([exp_biom[u] for u in units_in_loc])
            labels = [u.split(" :: ")[-1] if " :: " in u else u for u in units_in_loc]
            ax.stackplot(months, stack, labels=labels,
                         colors=[cmap(j % 10) for j in range(len(units_in_loc))], alpha=0.75)
            if loc in self.loc_mab:
                mab_t = self.loc_mab[loc] / 1e3
                ax.axhline(mab_t, color="red", ls="--", lw=1.5,
                           label=f"Location MAB  {mab_t:,.0f} t")
            ax.set_ylabel("Biomass (t)")
            ax.set_title(f"{loc}  (expected across {self.n_scenarios} scenarios)", fontsize=10)
            ax.legend(loc="upper left", fontsize=7, ncol=4, framealpha=0.6)
            ax.grid(True, alpha=0.25)

        ax = axes[-1]
        for sc in range(self.n_scenarios):
            sc_regional = sum(unit_biom[sc][u] for u in U)
            ax.plot(months, sc_regional, color="steelblue", lw=0.8, alpha=0.35)
        exp_regional = sum(exp_biom[u] for u in U)
        ax.fill_between(months, exp_regional, alpha=0.45, color="steelblue")
        ax.plot(months, exp_regional, color="steelblue", lw=2.0,
                label=f"Expected ({self.n_scenarios} scenarios)")
        reg_mab_t = self.regional_mab / 1e3
        ax.axhline(reg_mab_t, color="red", ls="--", lw=1.5,
                   label=f"Regional MAB  {reg_mab_t:,.0f} t")
        ax.axvline(self.stage0_months[-1] + 1, color="purple", ls="--", lw=0.6, alpha=0.3)
        ax.set_ylabel("Biomass (t)")
        ax.set_xlabel("Month")
        ax.set_title("Regional Total", fontsize=10)
        ax.legend(loc="upper left", fontsize=8, framealpha=0.6)
        ax.grid(True, alpha=0.25)

        fig.suptitle(
            f"Expected Biomass over Time — DE ({self.n_scenarios} scenarios)",
            fontsize=12, fontweight="bold",
        )
        fig.tight_layout()
        if output_file is not None:
            plt.savefig(output_file, dpi=150)
            print(f"Biomass plot saved to {output_file}")
        return fig

    def plot(self, output_file: str = "stats/de_solution.png"):
        """
        1×3 figure: Bad / Normal / Good path stocking & harvest timelines (60 months).
        """
        if self.model is None or self.model.SolCount == 0:
            print("No solution to plot.")
            return
        matplotlib.use("Agg")

        ref        = self.milp_objects[0]
        U          = ref.U
        U_exist    = ref.U_exist
        U_exist_set = set(U_exist)
        loc_of     = ref.loc_of
        loc_names  = sorted(set(loc_of.values()))
        palette    = ["tab:blue", "tab:orange", "tab:green",
                      "tab:purple", "tab:brown", "tab:pink"]
        loc_col    = {ln: palette[i] for i, ln in enumerate(loc_names)}
        unit_to_y  = {u: i for i, u in enumerate(U)}
        n_units    = len(U)
        T          = self.T
        BH         = 0.45
        s1_start   = self.stage1_months[0]

        def short_lbl(u):
            lp, up = u.split(" :: ")
            return f"L{lp.split()[-1]}:U{up.split()[-1]}"

        def _draw_timeline(ax, sc, title, shade_color):
            milp = self.milp_objects[sc]

            ax.axvline(s1_start, color="gray", ls="--", lw=0.8, alpha=0.6)
            ax.text(s1_start + 0.4, n_units - 0.1, "S1",
                    fontsize=6, color="gray", va="top")
            ax.axvspan(s1_start, T, alpha=0.07, color=shade_color, zorder=0)
            ax.text((s1_start + T) / 2, n_units - 0.15, "stage-1",
                    ha="center", va="top", fontsize=6,
                    color=shade_color, style="italic")

            for u in U:
                y   = unit_to_y[u]
                col = loc_col[loc_of[u]]

                if u in U_exist_set:
                    harvested = False
                    for t in range(T):
                        if self._h_exist[sc, u, t].X > 0.5:
                            ax.barh(y, t, left=0, height=BH,
                                    color=col, alpha=0.25, edgecolor="none", zorder=1)
                            ax.plot(t, y, "v", color="darkorange", ms=6, zorder=4)
                            harvested = True
                            break
                    if not harvested:
                        ax.barh(y, T, left=0, height=BH,
                                color=col, alpha=0.10, edgecolor="none", zorder=1)

                for s_start in milp.Tset:
                    if self._stk[sc, u, s_start].X < 0.5:
                        continue
                    ax.plot(s_start, y, "^", color="tab:green", ms=6, zorder=5)
                    for th in milp.H_by_us.get((u, s_start), []):
                        if self._harv[sc, u, s_start, th].X > 0.5:
                            ax.barh(y, th - s_start, left=s_start,
                                    height=BH, color=col, alpha=0.55,
                                    edgecolor=col, lw=0.5, zorder=2)
                            ax.plot(th, y, "v", color="tab:red", ms=5, zorder=4)

            ax.set_xlim(0, T)
            ax.set_ylim(-0.7, n_units - 0.3)
            ax.set_yticks(range(n_units))
            ax.set_yticklabels([short_lbl(u) for u in U], fontsize=7)
            ax.set_xlabel("Month")
            ax.set_title(title)
            ax.grid(True, axis="x", alpha=0.25)

            leg = [
                _L2D([0], [0], marker="^", color="w",
                     markerfacecolor="tab:green",  ms=7, label="Stocking"),
                _L2D([0], [0], marker="v", color="w",
                     markerfacecolor="tab:red",    ms=7, label="Harvest (new)"),
                _L2D([0], [0], marker="v", color="w",
                     markerfacecolor="darkorange", ms=7, label="Harvest (exist.)"),
            ] + [_MPatch(color=palette[i], alpha=0.55,
                         label=ln.replace("Location ", "Loc"))
                 for i, ln in enumerate(loc_names)]
            ax.legend(handles=leg, fontsize=6, loc="upper right", ncol=1)

        fig, axes = plt.subplots(2, 3, figsize=(21, 10))

        bad_sc    = next((i for i, nm in enumerate(self.scenario_names)
                          if nm == "s1_bad__s2_bad"),    0)
        normal_sc = next((i for i, nm in enumerate(self.scenario_names)
                          if nm == "s1_normal__s2_normal"), 4)
        good_sc   = next((i for i, nm in enumerate(self.scenario_names)
                          if nm == "s1_good__s2_good"), 8)

        _draw_timeline(axes[0, 0], bad_sc,    "Bad Path  (s1_bad → s2_bad)",       "tab:red")
        _draw_timeline(axes[0, 1], normal_sc, "Normal Path  (s1_normal → s2_normal)", "steelblue")
        _draw_timeline(axes[0, 2], good_sc,   "Good Path  (s1_good → s2_good)",    "tab:green")

        axes[1, 0].axis("off")
        axes[1, 1].axis("off")

        status_map = {2: "OPTIMAL", 3: "INFEASIBLE", 5: "UNBOUNDED", 9: "TIME_LIMIT"}
        axes[1, 2].text(
            0.5, 0.5,
            f"Deterministic Equivalent — Salmon Farming MILP\n\n"
            f"Scenarios  : {self.n_scenarios}  (3², 2 stages of uncertainty)\n"
            f"Status     : {status_map.get(self.status, str(self.status))}\n"
            f"Obj value  : {self.obj_val:,.0f} NOK\n"
            f"MIP gap    : {self.model.MIPGap:.4%}\n"
            f"Solve time : {self.solve_time:.1f}s\n\n"
            f"--- Scenario Tree ---\n"
            f"Stage-0: 1 node   × 9 scenarios\n"
            f"Stage-1: 3 nodes  × 3 scenarios",
            transform=axes[1, 2].transAxes,
            fontsize=11, va="center", ha="center",
            bbox=dict(boxstyle="round", facecolor="wheat", alpha=0.5),
        )
        axes[1, 2].set_title("Summary")
        axes[1, 2].axis("off")

        fig.suptitle("DE Solution — Stocking & Harvest Plans",
                     fontsize=13, fontweight="bold")
        fig.tight_layout()
        plt.savefig(output_file, dpi=150)
        print(f"\nPlot saved to {output_file}")


# ============================================================================
# export_decision_tree — Excel workbook showing the DE decision tree
# ============================================================================

def export_decision_tree(de: "DeterministicEquivalent",
                         filename: str = "de_decision_tree.xlsx"):
    """Export the DE decision tree to a readable Excel workbook."""
    if de.model is None or de.model.SolCount == 0:
        print("DE has no solution — cannot export decision tree.")
        return

    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        _have_openpyxl = True
    except ImportError:
        _have_openpyxl = False

    labels = de.stage_labels
    ref    = de.milp_objects[0]

    def _parse_sname(name):
        p = name.split("__")
        return (p[0].split("_", 1)[1], p[1].split("_", 1)[1])

    sc_nodes: list[dict] = []
    node_to_rep: dict[str, int] = {}
    node_to_scens: dict[str, list[int]] = {}

    for sc_idx, name in enumerate(de.scenario_names):
        t1, t2 = _parse_sname(name)
        nk0 = "s0"
        nk1 = f"s1_{t1}"
        sc_nodes.append({"s0": nk0, "s1": nk1})
        for nk in (nk0, nk1):
            node_to_scens.setdefault(nk, []).append(sc_idx)
            if nk not in node_to_rep:
                node_to_rep[nk] = sc_idx

    stage0_set = set(de.stage0_months)
    stage1_set = set(de.stage1_months)

    def _months_for_prefix(nk: str) -> set:
        if nk == "s0": return stage0_set
        return stage1_set

    def _active_actions(nk: str) -> list[dict]:
        sc     = node_to_rep[nk]
        months = _months_for_prefix(nk)
        acts   = []

        for u in ref.U:
            for t in sorted(months):
                if round(de._stk[sc, u, t].X) >= 1:
                    q_val = de._q[sc, u, t].X
                    acts.append({
                        "action": "Stock", "unit": u, "month": t,
                        "cohort_start": "", "qty_fish": f"{int(round(q_val)):,}",
                        "notes": f"Stock {u} in month {t} ({int(round(q_val)):,} fish)",
                    })

        for (sc2, u, s_start, th), var in de._harv.items():
            if sc2 != sc or th not in months:
                continue
            if round(var.X) >= 1:
                acts.append({
                    "action": "Harvest", "unit": u, "month": th,
                    "cohort_start": s_start, "qty_fish": "",
                    "notes": f"Harvest {u} in month {th}  (cohort stocked month {s_start})",
                })

        for u in ref.U_exist:
            for t in sorted(months):
                if round(de._h_exist[sc, u, t].X) >= 1:
                    acts.append({
                        "action": "Harvest (existing)", "unit": u, "month": t,
                        "cohort_start": "existing", "qty_fish": "",
                        "notes": f"Harvest existing stock in {u} in month {t}",
                    })

        acts.sort(key=lambda a: (0 if a["action"] == "Stock" else 1, a["month"], a["unit"]))
        return acts

    def _count_actions(nk: str, action: str) -> int:
        return sum(1 for a in _active_actions(nk) if a["action"].startswith(action))

    lookup_rows = []
    for sc_idx, name in enumerate(de.scenario_names):
        t1, t2 = _parse_sname(name)
        nk1 = sc_nodes[sc_idx]["s1"]
        lookup_rows.append({
            "scenario_id":  sc_idx,
            "t1 (stage0)":  t1,
            "t2 (stage1)":  t2,
            "stage1_node":  nk1,
            "s0_stocks":    _count_actions("s0", "Stock"),
            "s0_harvests":  _count_actions("s0", "Harvest"),
            "s1_stocks":    _count_actions(nk1, "Stock"),
            "s1_harvests":  _count_actions(nk1, "Harvest"),
            "feasible":     "YES",
            "obj_value":    round(de.obj_val, 0),
        })
    df_lookup = pd.DataFrame(lookup_rows)

    def _stage_from_month(m: int) -> str:
        if m in stage0_set: return "Stage0"
        return "Stage1"

    _stage_order = {"Stage0": 0, "Stage1": 1}

    playbook_rows = []
    for sc_idx, name in enumerate(de.scenario_names):
        t1, t2 = _parse_sname(name)

        for u in ref.U:
            for t in ref.Tset:
                if round(de._stk[sc_idx, u, t].X) >= 1:
                    q_val = de._q[sc_idx, u, t].X
                    stage = _stage_from_month(t)
                    playbook_rows.append({
                        "scenario_id": sc_idx, "t1": t1, "t2": t2,
                        "stage": stage, "stage_order": _stage_order[stage],
                        "node": sc_nodes[sc_idx][f"s{_stage_order[stage]}"],
                        "action": "Stock", "unit": u, "action_month": t,
                        "cohort_start": "", "qty_fish": int(round(q_val)),
                    })

        for (sc2, u, s_start, th), var in de._harv.items():
            if sc2 != sc_idx:
                continue
            if round(var.X) >= 1:
                stage = _stage_from_month(th)
                playbook_rows.append({
                    "scenario_id": sc_idx, "t1": t1, "t2": t2,
                    "stage": stage, "stage_order": _stage_order[stage],
                    "node": sc_nodes[sc_idx][f"s{_stage_order[stage]}"],
                    "action": "Harvest", "unit": u, "action_month": th,
                    "cohort_start": s_start, "qty_fish": "",
                })

        for u in ref.U_exist:
            for t in ref.Tset:
                if round(de._h_exist[sc_idx, u, t].X) >= 1:
                    stage = _stage_from_month(t)
                    playbook_rows.append({
                        "scenario_id": sc_idx, "t1": t1, "t2": t2,
                        "stage": stage, "stage_order": _stage_order[stage],
                        "node": sc_nodes[sc_idx][f"s{_stage_order[stage]}"],
                        "action": "Harvest (existing)", "unit": u, "action_month": t,
                        "cohort_start": "existing", "qty_fish": "",
                    })

    df_playbook = (pd.DataFrame(playbook_rows)
                   .sort_values(["scenario_id", "stage_order", "action_month", "unit"])
                   .reset_index(drop=True)
                   if playbook_rows else pd.DataFrame())

    guide_rows = [
        {"Sheet": "Tree_Plan",
         "What it shows": "The complete 4-node decision tree. "
                          "Stage 0 (1 node) → Stage 1 (3 nodes). "
                          "Stage-0 decisions are committed before any uncertainty resolves.",
         "How to use": "Read Stage 0 first. After month 29, observe t1 and jump to your Stage-1 branch."},
        {"Sheet": "Scenario_Lookup",
         "What it shows": "9 rows — one per scenario with node keys and decision counts.",
         "How to use": "Find your scenario by t1/t2 to identify relevant Tree_Plan sections."},
        {"Sheet": "Scenario_Playbook",
         "What it shows": "Full chronological action list for every scenario across both stages.",
         "How to use": "Filter by scenario_id to see the exact sequence of Stock / Harvest actions."},
    ]
    df_guide = pd.DataFrame(guide_rows)

    if not _have_openpyxl:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_guide.to_excel(writer, sheet_name="Guide", index=False)
            df_lookup.to_excel(writer, sheet_name="Scenario_Lookup", index=False)
            df_playbook.to_excel(writer, sheet_name="Scenario_Playbook", index=False)
        print(f"Decision tree exported to {filename} (no rich formatting — install openpyxl)")
        return

    C_STAGE   = PatternFill("solid", fgColor="1F4E79")
    C_NODE    = PatternFill("solid", fgColor="2E75B6")
    C_COL_HDR = PatternFill("solid", fgColor="D6E4F0")
    C_STOCK   = PatternFill("solid", fgColor="E2EFDA")
    C_HARV    = PatternFill("solid", fgColor="FCE4D6")
    C_NONE    = PatternFill("solid", fgColor="F2F2F2")
    C_BLUE_H  = PatternFill("solid", fgColor="BDD7EE")
    C_GREY_H  = PatternFill("solid", fgColor="D9D9D9")

    F_WHITE_BOLD = Font(bold=True, color="FFFFFF", size=11)
    F_WHITE      = Font(color="FFFFFF", italic=True)
    F_BOLD       = Font(bold=True)
    F_NORMAL     = Font()

    THIN = Side(style="thin", color="AAAAAA")
    BORD = Border(left=THIN, right=THIN, bottom=THIN, top=THIN)

    TP_COLS = ["Action", "Unit", "Month", "Cohort_Start", "Qty_Fish", "Notes"]
    NCOLS   = len(TP_COLS)

    wb    = openpyxl.Workbook()
    ws_tp = wb.active
    ws_tp.title = "Tree_Plan"
    ws_tp.freeze_panes = "A3"

    col_widths = [22, 20, 9, 14, 14, 50]
    for ci, w in enumerate(col_widths, 1):
        ws_tp.column_dimensions[get_column_letter(ci)].width = w

    row = 1

    def _fill_row(ws, r, fill, font=None, values=None):
        for c in range(1, NCOLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            if font:
                cell.font = font
            cell.border = BORD
            if values and (c - 1) < len(values):
                cell.value = values[c - 1]

    def _write_stage_banner(ws, r, title, subtitle):
        _fill_row(ws, r, C_STAGE, F_WHITE_BOLD)
        ws.cell(row=r, column=1).value = title
        ws.cell(row=r, column=2).value = subtitle
        ws.cell(row=r, column=2).font  = F_WHITE
        ws.row_dimensions[r].height = 18

    def _write_node_banner(ws, r, title, subtitle):
        _fill_row(ws, r, C_NODE, F_WHITE_BOLD)
        ws.cell(row=r, column=1).value = "  " + title
        ws.cell(row=r, column=2).value = subtitle
        ws.cell(row=r, column=2).font  = Font(color="FFFFFF", italic=True)
        ws.row_dimensions[r].height = 16

    def _write_col_headers(ws, r):
        _fill_row(ws, r, C_COL_HDR, F_BOLD, TP_COLS)

    def _write_action_row(ws, r, action, unit, month, cohort_start, qty, notes):
        fill = C_STOCK if action == "Stock" else C_HARV
        _fill_row(ws, r, fill, F_NORMAL,
                  [action, unit, month, cohort_start,
                   int(qty) if isinstance(qty, (int, float)) and qty != "" else (qty or ""),
                   notes])

    def _write_no_decisions(ws, r):
        _fill_row(ws, r, C_NONE)
        ws.cell(row=r, column=1).value = "  (no decisions in this node)"
        ws.cell(row=r, column=1).font  = Font(italic=True, color="888888")

    _write_stage_banner(ws_tp, row,
                        "STAGE 0  —  Months 0–29",
                        "Committed before any uncertainty resolves.  Identical for ALL 9 scenarios.")
    row += 1
    _write_col_headers(ws_tp, row); row += 1
    for act in _active_actions("s0"):
        _write_action_row(ws_tp, row, act["action"], act["unit"],
                          act["month"], act["cohort_start"],
                          act["qty_fish"], act["notes"])
        row += 1
    if not _active_actions("s0"):
        _write_no_decisions(ws_tp, row); row += 1
    row += 1

    _write_stage_banner(ws_tp, row,
                        "STAGE 1  —  Months 30–59",
                        "Decided after observing t1.  Three branches.")
    row += 1
    for t1 in labels:
        nk      = f"s1_{t1}"
        n_scens = len(node_to_scens.get(nk, []))
        _write_node_banner(ws_tp, row,
                           f"IF  t1 = {t1.upper()}",
                           f"{n_scens} of 9 scenarios follow this branch")
        row += 1
        _write_col_headers(ws_tp, row); row += 1
        actions = _active_actions(nk)
        for act in actions:
            _write_action_row(ws_tp, row, act["action"], act["unit"],
                              act["month"], act["cohort_start"],
                              act["qty_fish"], act["notes"])
            row += 1
        if not actions:
            _write_no_decisions(ws_tp, row); row += 1
        row += 1

    def _add_df_sheet(wb, df, sheet_name, header_fill):
        ws = wb.create_sheet(title=sheet_name)
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = F_BOLD
            cell.border = BORD
        for ri, row_data in enumerate(df.itertuples(index=False), start=2):
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ri, column=ci,
                               value=(None if (isinstance(val, float) and np.isnan(val)) else val))
                cell.border = BORD
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0) for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)
        ws.freeze_panes = "A2"
        return ws

    _add_df_sheet(wb, df_lookup,   "Scenario_Lookup",   C_BLUE_H)
    _add_df_sheet(wb, df_playbook, "Scenario_Playbook", C_BLUE_H)

    ws_g = wb.create_sheet(title="Guide")
    for ci, col in enumerate(df_guide.columns, 1):
        cell = ws_g.cell(row=1, column=ci, value=col)
        cell.fill = C_GREY_H; cell.font = F_BOLD; cell.border = BORD
    for ri, row_data in enumerate(df_guide.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, start=1):
            cell = ws_g.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORD
    ws_g.column_dimensions["A"].width = 20
    ws_g.column_dimensions["B"].width = 60
    ws_g.column_dimensions["C"].width = 60
    for r in range(2, len(df_guide) + 2):
        ws_g.row_dimensions[r].height = 55

    wb.save(filename)
    print(f"Decision tree exported to {filename}")
    print(f"  Sheets: Tree_Plan | Scenario_Lookup | Scenario_Playbook | Guide")


# ============================================================================
# ENTRY POINT
# ============================================================================
if __name__ == "__main__":
    from instance import units_df, loc_mab, regional_mab, T, temps_normal_12, temps_bad_12, temps_good_12

    de = DeterministicEquivalent(
        units_df=units_df,
        loc_mab=loc_mab,
        regional_mab=regional_mab,
        T=T,
        temps_bad=temps_bad_12,
        temps_normal=temps_normal_12,
        temps_good=temps_good_12,
        mip_gap=0.02,
        seed=None,
    )
    de.build()
    de.solve()
    de.print_results()
    # export_decision_tree(de, filename="stats/de_decision_tree.xlsx")
    # de.plot("stats/de_solution.png")
    # de.plot_biomass("stats/de_biomass.png")
    status_map = {2: "OPTIMAL", 3: "INFEASIBLE", 5: "UNBOUNDED", 9: "TIME_LIMIT"}
    print("\n" + "=" * 70)
    print(f"Status     : {status_map.get(de.status, str(de.status))}")
    print(f"Solve time : {de.solve_time:.1f}s")
    print(f"Obj value  : {de.obj_val:,.2f}")
    print(f"MIP gap    : {de.model.MIPGap:.4%}")
    print("=" * 70)
