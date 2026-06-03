"""
Augmented Lagrangian Decomposition for Salmon Farming MILP — tree-aware NA version.

  SP.py (tree-aware):
    - Stage-0 variables are shared across all 81 scenarios         ("z[u,0]@s0")
    - Stage-1 variables are shared within the 3×27 branches        ("z[u,15]@s1_bad" etc.)
    - Stage-2 variables are shared within the 9×9 branches         ("z[u,30]@s2_bad_bad" etc.)
    - Stage-3 variables are shared within the 27×3 branches        ("z[u,45]@s3_bad_bad_bad" etc.)
    - x̄ for each qualified variable is computed as the conditional
      probability-weighted average over scenarios in that node only.
    - Multiplier and fixing logic is similarly node-scoped.

Key changes versus SP.py
─────────────────────────
1. _build_variable_index  — corrected node labels; @node suffix on all keys;
                             weight matrix _xbar_W; participating mask _participating.
2. _compute_xbar          — (W * sol_matrix).sum(axis=0) instead of probs @ sol_matrix.
3. Variable fixing loop   — per-node consensus check instead of all-scenarios check.
4. Convergence / devs     — participating mask applied before computing deviations.
5. Multiplier update      — participating mask zeroes out non-participating (s,vi).
6. Slam repair            — strips @node suffix when parsing variable names.
7. _post_process          — strips @node suffix when parsing variable names.
8. plot()                 — strips @node suffix when parsing variable names.
"""

import os
import itertools
from collections import defaultdict
from concurrent.futures import FIRST_COMPLETED, ThreadPoolExecutor, as_completed, wait
from time import time

from tqdm import tqdm

import matplotlib
matplotlib.use("Agg")
import matplotlib.pyplot as plt
import numpy as np
import pandas as pd
import gurobipy as gb
from gurobipy import GRB

from ip import SalmonFarmingMILP
from instance import units_df, loc_mab, regional_mab


class BinaryProgressiveHedging:
    """
    Tree-aware Augmented Lagrangian Decomposition for the 81-scenario 4-stage
    salmon farming MILP.

    Non-anticipativity is enforced per node of the scenario tree:
      - Stage-0 vars (@s0)              : same across all 81 scenarios
      - Stage-1 vars (@s1_{t1})         : same within 27-scenario stage-1 branches
      - Stage-2 vars (@s2_{t1}_{t2})    : same within 9-scenario stage-1+2 branches
      - Stage-3 vars (@s3_{t1}_{t2}_{t3}): same within 3-scenario stage-1+2+3 branches
    """

    BINARY_PREFIXES = {"z", "h", "h_exist"}

    def __init__(
        self,
        units_df: pd.DataFrame,
        loc_mab: dict,
        regional_mab: float,
        *,
        T: int = 60,
        K: int = 1_000,
        epsilon_bin: float = 0.01,
        epsilon_cont: float = 0.05,
        rho_bin: float = 50.0,
        rho_cont: float = 3.0,
        penalty_fraction: float = 0.1,
        rho_increase: float = 1.3,
        rho_max_bin: float = 1e6,
        rho_max_cont: float = 1e4,
        rho_max_obj_fraction: float = 10.0,
        stall_window: int = 3,
        slam_stall_window: int = 6,
        slam_stall_tol: float = 0.01,
        min_iters_before_slam: int = 10,
        slam_max_dev_threshold: float = 0.5,
        tail_fix_window: int = 5,
        tail_fix_max_disagree: int = 200,
        fix_threshold: int = 3,
        mip_gap: float = 0.00,
        temps_normal: np.ndarray = None,
        temps_bad: np.ndarray = None,
        temps_good: np.ndarray = None,
        S_normal: float = None,
        S_bad: float = None,
        show_subproblem_incumbents: bool = False,
        show_subproblem_mip_progress: bool = False,
        # ─── Rolling-horizon parameters ────────────────────────────────────
        # If `prefix_months` is non-empty, all months in it use a single
        # deterministic temperature label (`prefix_label`). The first stochastic
        # stage is then "blind" — its decisions share the same NA group ("s0")
        # as the prefix. Total scenarios = 3^len(stage_slices).
        # Defaults preserve the original 60-month, 81-scenario behaviour
        # (no prefix, four equal 15-month stages).
        stage_slices: list = None,
        prefix_months: list = None,
        prefix_label:  str  = None,
        start_calendar_month: int = 0,
    ):
        self.units_df         = units_df
        self.loc_mab          = loc_mab
        self.regional_mab     = regional_mab
        self.T                = T
        self.K                = K
        self.epsilon_bin      = epsilon_bin
        self.epsilon_cont     = epsilon_cont
        self.rho_bin          = rho_bin
        self.rho_cont         = rho_cont
        self.penalty_fraction = penalty_fraction
        self.rho_increase     = rho_increase
        self.rho_max_bin          = rho_max_bin
        self.rho_max_cont         = rho_max_cont
        self.rho_max_obj_fraction = rho_max_obj_fraction
        self.stall_window          = stall_window
        self.slam_stall_window      = slam_stall_window
        self.slam_stall_tol         = slam_stall_tol
        self.min_iters_before_slam  = min_iters_before_slam
        self.slam_max_dev_threshold = slam_max_dev_threshold
        self.tail_fix_window        = tail_fix_window
        self.tail_fix_max_disagree  = tail_fix_max_disagree
        self.fix_threshold         = fix_threshold
        self.mip_gap          = mip_gap
        self.show_subproblem_incumbents = show_subproblem_incumbents
        self.show_subproblem_mip_progress = show_subproblem_mip_progress

        # ─── Stage / prefix configuration ─────────────────────────────────
        if stage_slices is None:
            # Original default: 4 evenly spaced stages over T months
            stage_len = T // 4
            stage_slices = [
                list(range(0,             stage_len)),
                list(range(stage_len,     2 * stage_len)),
                list(range(2 * stage_len, 3 * stage_len)),
                list(range(3 * stage_len, T)),
            ]
        if prefix_months is None:
            prefix_months = []

        self.stage_slices         = [list(sl) for sl in stage_slices]
        self.prefix_months        = list(prefix_months)
        self.prefix_label         = prefix_label
        self.start_calendar_month = int(start_calendar_month) % 12
        self.n_branching_stages   = len(self.stage_slices)

        # Backward-compatible stage{0..3}_months / _end attributes.
        # The first stochastic stage shares the "s0" NA group with the
        # deterministic prefix when prefix_months is non-empty, so we treat
        # (prefix + first stage) as the single "stage 0" block for the
        # if/elif chains in _post_process / consistency repair.
        if self.prefix_months:
            stage0_block = sorted(self.prefix_months + self.stage_slices[0])
            stage_blocks = [stage0_block] + [list(sl) for sl in self.stage_slices[1:]]
        else:
            stage_blocks = [list(sl) for sl in self.stage_slices]

        # Pad / truncate stage_blocks to length 4 for backward compat
        while len(stage_blocks) < 4:
            # Empty trailing stage(s); _end equals previous _end.
            stage_blocks.append([])

        self.stage0_months = stage_blocks[0]
        self.stage1_months = stage_blocks[1]
        self.stage2_months = stage_blocks[2]
        self.stage3_months = stage_blocks[3]

        def _end(months, prev_end):
            return (max(months) + 1) if months else prev_end

        self.stage0_end = _end(self.stage0_months, 0)
        self.stage1_end = _end(self.stage1_months, self.stage0_end)
        self.stage2_end = _end(self.stage2_months, self.stage1_end)
        self.stage3_end = _end(self.stage3_months, self.stage2_end)

        # Map month → branching-stage index (-1 for prefix months which share
        # the same NA group "s0" as the first stochastic stage). This is the
        # source of truth used by _node_key_for / _build_variable_index to
        # construct NA node labels in a way that works for any number of
        # stages (with or without a deterministic prefix).
        self._month_to_stage: list = [None] * T
        for m in self.prefix_months:
            if 0 <= m < T:
                self._month_to_stage[m] = -1  # prefix → "s0"
        for k, sl in enumerate(self.stage_slices):
            for m in sl:
                if 0 <= m < T:
                    self._month_to_stage[m] = k

        # Temperature patterns (matching DE.py)
        days = 30
        # self.temps_normal = (temps_normal if temps_normal is not None
        #                      else np.array([5, 5, 5, 6, 9, 12, 14, 16.5, 15.5, 13, 10, 7.5]))
        self.temps_normal = (temps_normal if temps_normal is not None
                             else np.array([5, 5, 5, 6, 9, 12, 14, 16.5, 15.5, 13, 10, 7.5]))
        # self.temps_bad    = (temps_bad    if temps_bad    is not None
        #                      else np.array([3, 3, 3, 4, 7, 10, 12, 14.5, 13.5, 11,  8, 5.5]))
        self.temps_bad    = (temps_bad    if temps_bad    is not None
                             else np.array([3, 3, 3, 4,  7, 10, 12, 14.5, 13.5, 11,  8, 5.5]))
        self.temps_good   = (temps_good   if temps_good   is not None
                             else np.array([7, 7, 7, 8, 11, 14, 16, 18.5, 17.5, 15, 12, 9.5]))
        # self.temps_good   = (temps_good   if temps_good   is not None
        #                      else np.array([7, 7, 7, 8, 11, 14, 16, 18.5, 17.5, 15, 12, 9.5]))
        self.S_normal     = (1.0 - 0.0002) ** days
        self.S_bad        = (1.0 - 0.001)  ** days # altered
        self.labels       = ["bad", "normal", "good"]

        # Populated by build()
        self.scenarios:                  list       = []
        self.milp_objects:               list       = []
        self.probabilities:              list       = []
        self.scenario_names:             list       = []
        self.bundle_id:                  list       = []
        self.n_scenarios:                int        = 0
        self.probs:                      np.ndarray = None
        self.all_na_names_ordered:       list       = []  # qualified names (with @node)
        self.name_to_idx:                dict       = {}
        self.V:                          int        = 0
        self.is_binary:                  np.ndarray = None
        self.n_binary_na:                int        = 0
        self.n_cont_na:                  int        = 0
        self.scenario_vidx:              list       = []
        self.var_cache_flat:             list       = []
        self.na_var_names_per_scenario:  list       = []
        self.original_objectives:        list       = []
        self.sol_matrix:                 np.ndarray = None
        self.multipliers:                np.ndarray = None
        self.var_scale:                  np.ndarray = None
        self.fixed:                      np.ndarray = None
        self.fix_val:                    np.ndarray = None
        self.unanimous_count:            np.ndarray = None
        self.xbar_binary_history:        dict       = {}
        self.x_bar:                      np.ndarray = None
        self.obj_lin:                    list       = []
        # Tree-aware additions
        self._xbar_W:                    np.ndarray = None  # weight matrix (n_scen, V)
        self._participating:             np.ndarray = None  # bool mask (n_scen, V)
        self._node_scens:                dict       = {}    # node_key -> list[int]
        self._vi_to_node:                list       = []    # vi -> node_key
        self._vi_occurrences:            list       = []    # vi -> list[(scenario_idx, local_var_idx)]

        # Results — populated by solve()
        self.converged:            bool  = False
        self.convergence_history:  list  = []
        self.objective_history:    list  = []
        self.n_fixed_history:      list  = []
        self.avg_dev_bin_history:  list  = []
        self.avg_dev_cont_history: list  = []
        self.total_time:           float = 0.0
        self.n_iters:              int   = 0
        self.max_dev_bin:          float = 0.0
        self.avg_dev_bin:          float = 0.0
        self.max_dev_cont:         float = 0.0
        self.avg_dev_cont:         float = 0.0

        # Post-processing results — populated by solve()
        self.eval_obj:             float = 0.0
        self.n_feasible:           int   = 0
        self.n_infeasible:         int   = 0
        self.infeasible_scenarios: list  = []
        self.bin_consensus:        dict  = {}
        self._stock_na:            dict  = {}
        self._harv_na:             dict  = {}
        self._hexist_na:           dict  = {}
        self.q_robust_vals:        dict  = {}


    # =========================================================================
    # Static helpers
    # =========================================================================

    @staticmethod
    def _tile(monthly_12: np.ndarray, n: int) -> np.ndarray:
        return np.tile(monthly_12, (n // 12) + 1)[:n]

    @staticmethod
    def _stage_surv(label: str, prev_label: str, S_normal: float, S_bad: float) -> float:
        """High mortality only when this AND the preceding stage are both 'good'."""
        return S_bad if (label == "good" and prev_label == "good") else S_normal
        

    @classmethod
    def _is_binary_name(cls, vname: str) -> bool:
        return vname.split("[")[0] in cls.BINARY_PREFIXES

    def _node_key_for(self, s: int, m: int) -> str:
        """NA tree node key for scenario `s` at month `m`.

        - Prefix months and the first stochastic stage (k=0) share "s0" — they
          are decided before any temperature is observed.
        - Stage k≥1 uses f"s{k}_<labels of stages 0..k-1>"; the labels of
          earlier stages are the information available at decision time.
        """
        k = self._month_to_stage[m]
        if k is None or k <= 0:
            return "s0"
        label_tuple = self._label_tuples[s]
        return f"s{k}_" + "_".join(label_tuple[:k])

    # =========================================================================
    # build() — build scenario models and variable index
    # =========================================================================

    def _calendar_temps(self, label: str, months: list, temp_map: dict) -> np.ndarray:
        """Calendar-aware monthly temperatures: horizon-relative t=0 maps to
        calendar month `self.start_calendar_month`, then wraps around the
        12-month seasonal profile for `label`."""
        profile = temp_map[label]
        off = self.start_calendar_month % 12
        return np.array([profile[(off + m) % 12] for m in months])

    def build(self):
        """Build scenario models and the NA variable indexing structures.

        Total scenarios = 3^n_branching_stages.
        - 4 branching stages, no prefix → 81 scenarios (original behavior).
        - 3 branching stages with 3-month prefix → 27 scenarios (rolling mode).
        """
        T       = self.T
        labels  = self.labels
        n_branching = self.n_branching_stages
        n_total = len(labels) ** n_branching
        temp_map = {
            "bad":    self.temps_bad,
            "normal": self.temps_normal,
            "good":   self.temps_good,
        }

        # Pre-compute the canonical scenario ordering: itertools.product over
        # n_branching stages. Scenario s maps to label_tuple_for_scenario[s].
        self._label_tuples = list(itertools.product(labels, repeat=n_branching))

        scenarios, milp_objects, probabilities, scenario_names, bundle_id = [], [], [], [], []

        for label_tuple in self._label_tuples:
            # Build name: include prefix tag if a deterministic prefix is used.
            tag_parts = [f"s{i+1}_{l}" for i, l in enumerate(label_tuple)]
            if self.prefix_months:
                name = f"prefix_{self.prefix_label}__" + "__".join(tag_parts)
            else:
                name = "__".join(tag_parts)

            temps_t = np.zeros(T)
            S_t     = np.full(T, self.S_normal)

            # Block sequence: optional deterministic prefix, then stochastic stages.
            blocks = []
            if self.prefix_months:
                blocks.append((self.prefix_label, list(self.prefix_months)))
            for i, lbl in enumerate(label_tuple):
                blocks.append((lbl, self.stage_slices[i]))

            for i, (sl, months) in enumerate(blocks):
                prev = blocks[i - 1][0] if i > 0 else None
                surv = self._stage_surv(sl, prev, self.S_normal, self.S_bad)
                for m in months:
                    S_t[m] = surv
                cal_temps = self._calendar_temps(sl, months, temp_map)
                for m, val in zip(months, cal_temps):
                    temps_t[m] = val

            milp = SalmonFarmingMILP(
                units_df=self.units_df,
                temps_t=temps_t,
                survival_rates=S_t,
                loc_mab=self.loc_mab,
                regional_mab=self.regional_mab,
                horizon_months=T,
                scenario_name=name,
            )
            scenarios.append(milp.model)
            milp_objects.append(milp)
            probabilities.append(1.0 / n_total)
            scenario_names.append(name)
            # bundle by first stochastic label so parallel groups stay similar to before
            bundle_id.append(labels.index(label_tuple[0]))

        self.scenarios      = scenarios
        self.milp_objects   = milp_objects
        self.probabilities  = probabilities
        self.scenario_names = scenario_names
        self.bundle_id      = bundle_id
        self.n_scenarios    = len(scenarios)
        self.probs          = np.array(probabilities, dtype=np.float64)

        print(f"Built {self.n_scenarios} scenarios "
              f"({n_branching} stochastic stage(s)"
              + (f" + {len(self.prefix_months)}-month prefix" if self.prefix_months else "")
              + ")")

        # Build flat variable index and per-scenario caches
        self._build_variable_index()

        # Parallel setup
        import threading
        self._n_workers_actual = min(self.n_scenarios, os.cpu_count() or 4)
        self._progress_lock    = threading.Lock()
        self._progress_counter = [0]

        # Cache original (penalty-free) objectives before any setObjective calls
        self.original_objectives = [scenarios[s].getObjective() for s in range(self.n_scenarios)]

    def _build_variable_index(self):
        """
        Build NA variable name → global index mapping with correct non-anticipativity.

        Each variable is qualified as "varname@node" where node encodes only the
        information available at that decision stage. The tree generalises to any
        number of branching stages (with or without a deterministic prefix):

          - Prefix months & first stochastic stage (k=0): "@s0"
              (one node, all scenarios share one x̄)
          - Stage k≥1: "@s{k}_{l_0}_..._{l_{k-1}}"
              (3^k nodes, 3^(n_branching-k) scenarios each)
        """
        n_scenarios = self.n_scenarios
        n_branching = self.n_branching_stages

        na_var_names_per_scenario: list[list[str]] = []  # qualified names per scenario

        for s, milp in enumerate(self.milp_objects):
            label_tuple = self._label_tuples[s]

            node_var_pairs: list[tuple[str, str]] = []

            def _add_stage(node, months, _milp=milp):
                mset = set(months)
                for u in _milp.U:
                    for m in months:
                        node_var_pairs.append((node, f"z[{u},{m}]"))
                for u in _milp.U:
                    for s_start in _milp.Tset:
                        for t_harv in _milp.H_by_us.get((u, s_start), []):
                            if t_harv in mset:
                                node_var_pairs.append((node, f"h[{u},{s_start},{t_harv}]"))
                for u in _milp.U_exist:
                    for m in months:
                        node_var_pairs.append((node, f"h_exist[{u},{m}]"))

            # Deterministic prefix (if any) shares NA group "s0" with stage 0.
            if self.prefix_months:
                _add_stage("s0", self.prefix_months)
            for k in range(n_branching):
                if k == 0:
                    node_key = "s0"
                else:
                    node_key = f"s{k}_" + "_".join(label_tuple[:k])
                _add_stage(node_key, self.stage_slices[k])

            # Qualified names: "varname@node"
            qnames = [f"{vn}@{node}" for (node, vn) in node_var_pairs]
            na_var_names_per_scenario.append(qnames)

        # Global qualified name → index (dedup preserves first-seen order)
        all_na_names_ordered: list[str] = []
        name_to_idx: dict[str, int] = {}
        for s in range(n_scenarios):
            for qn in na_var_names_per_scenario[s]:
                if qn not in name_to_idx:
                    name_to_idx[qn] = len(all_na_names_ordered)
                    all_na_names_ordered.append(qn)

        V           = len(all_na_names_ordered)
        # Strip @node to get the actual variable name for is_binary check
        is_binary   = np.array(
            [self._is_binary_name(qn.split("@")[0]) for qn in all_na_names_ordered],
            dtype=bool
        )
        n_binary_na = int(is_binary.sum())
        n_cont_na   = V - n_binary_na
        print(f"NA variables: {V} total ({n_binary_na} binary, {n_cont_na} continuous)")

        # Per-scenario index arrays and Gurobi variable caches
        scenario_vidx:  list[np.ndarray] = []
        var_cache_flat: list[list]       = []
        vi_occurrences: list[list[tuple[int, int]]] = [[] for _ in range(V)]
        for s in range(n_scenarios):
            model = self.scenarios[s]
            model.update()
            idxs, vars_s = [], []
            for qn in na_var_names_per_scenario[s]:
                vi = name_to_idx[qn]
                vn = qn.split("@")[0]  # actual Gurobi variable name (no @node)
                local_i = len(idxs)
                idxs.append(vi)
                vars_s.append(model.getVarByName(vn))
                vi_occurrences[vi].append((s, local_i))
            scenario_vidx.append(np.array(idxs, dtype=np.int32))
            var_cache_flat.append(vars_s)

        print("Variable cache built (tree-aware)")

        # ── Node → scenario list ──────────────────────────────────────────────
        _node_scens: dict[str, list[int]] = defaultdict(list)
        for s in range(n_scenarios):
            label_tuple = self._label_tuples[s]
            # "s0" covers prefix (if any) + first stochastic stage.
            _node_scens["s0"].append(s)
            for k in range(1, n_branching):
                _node_scens[f"s{k}_" + "_".join(label_tuple[:k])].append(s)
        _node_scens = dict(_node_scens)

        # vi → node name
        _vi_to_node: list[str] = [qn.split("@")[1] for qn in all_na_names_ordered]

        # ── Node-aware weight matrix W[s, vi] ────────────────────────────────
        # W[s, vi] = conditional probability of scenario s within variable vi's node.
        # All scenarios are equally probable (1/81), so conditional prob = 1/|node|.
        W = np.zeros((n_scenarios, V), dtype=np.float64)
        for vi in range(V):
            node  = _vi_to_node[vi]
            scens = _node_scens[node]
            w     = 1.0 / len(scens)
            for s_n in scens:
                W[s_n, vi] = w

        # ── Participating mask _participating[s, vi] ─────────────────────────
        # True iff scenario s has variable vi in its NA set (i.e. is in vi's node).
        # Needed to avoid spurious deviations and multiplier updates for scenarios
        # that do not participate in a given node variable.
        _participating = np.zeros((n_scenarios, V), dtype=bool)
        for s in range(n_scenarios):
            for vi in scenario_vidx[s]:
                _participating[s, vi] = True

        self.na_var_names_per_scenario = na_var_names_per_scenario
        self.all_na_names_ordered      = all_na_names_ordered
        self.name_to_idx               = name_to_idx
        self.V                         = V
        self.is_binary                 = is_binary
        self.n_binary_na               = n_binary_na
        self.n_cont_na                 = n_cont_na
        self.scenario_vidx             = scenario_vidx
        self.var_cache_flat            = var_cache_flat
        self._node_scens               = _node_scens
        self._vi_to_node               = _vi_to_node
        self._vi_occurrences           = vi_occurrences
        self._xbar_W                   = W
        self._participating            = _participating

        # State arrays
        self.sol_matrix      = np.zeros((n_scenarios, V), dtype=np.float64)
        self.multipliers     = np.zeros((n_scenarios, V), dtype=np.float64)
        self.var_scale       = np.ones(V, dtype=np.float64)
        self.fixed           = np.zeros(V, dtype=bool)
        self.fix_val         = np.zeros(V, dtype=np.float64)
        self.unanimous_count = np.zeros(V, dtype=np.int32)
        self.xbar_binary_history = {v: [] for v in range(V) if is_binary[v]}

    # =========================================================================
    # Augmented objective infrastructure
    # =========================================================================

    def _compute_xbar(self) -> np.ndarray:
        """Node-aware x̄: weighted average within each variable's node."""
        return (self._xbar_W * self.sol_matrix).sum(axis=0)

    def _rho_eff_cont(self) -> np.ndarray:
        out = np.zeros(self.V, dtype=np.float64)
        out[~self.is_binary] = self.rho_cont / (self.var_scale[~self.is_binary] ** 2)
        return out

    def _build_augmented_objectives(self):
        """Set augmented objectives with quadratic penalty (called after rho calibration)."""
        rho_eff = self._rho_eff_cont()
        self.obj_lin = [
            np.zeros(len(self.scenario_vidx[s]), dtype=np.float64)
            for s in range(self.n_scenarios)
        ]
        for s in range(self.n_scenarios):
            model  = self.scenarios[s]
            vidx   = self.scenario_vidx[s]
            vars_s = self.var_cache_flat[s]
            orig   = self.original_objectives[s]

            quad = gb.QuadExpr()
            quad += orig
            for _, (vi, var) in enumerate(zip(vidx, vars_s)):
                if var is None:
                    continue
                if not self.is_binary[vi]:
                    quad.addTerms(rho_eff[vi] / 2.0, var, var)   # (ρ/2)x² — fixed
            model.setObjective(quad, GRB.MAXIMIZE)
            model.update()

        print("Augmented objectives built (quadratic part fixed, linear part updated per iter)")

    def _update_obj_linear(self, s: int, x_bar_vec: np.ndarray, lam_matrix: np.ndarray,
                           rho_b: float, rho_eff_vec: np.ndarray):
        """
        Update linear coefficients of the augmented objective for scenario s.

        Binary:     new_coeff = -(λ + (ρ/2)(1 - 2x̄))
        Continuous: new_coeff = -(λ - ρ_eff · x̄)

        x_bar_vec[vi] is now the node-specific x̄ for variable vi, so this
        method is unchanged — it simply uses the correctly computed x̄.
        """
        vidx      = self.scenario_vidx[s]
        vars_s    = self.var_cache_flat[s]
        old       = self.obj_lin[s]
        model     = self.scenarios[s]
        is_binary = self.is_binary
        fixed     = self.fixed

        changed_vars   = []
        changed_coeffs = []

        for local_i, (vi, var) in enumerate(zip(vidx, vars_s)):
            if var is None or fixed[vi]:
                continue
            xb  = x_bar_vec[vi]
            lam = lam_matrix[s, vi]
            if is_binary[vi]:
                new_coeff = -(lam + (rho_b / 2.0) * (1.0 - 2.0 * xb))
            else:
                new_coeff = -(lam - rho_eff_vec[vi] * xb)
            if abs(new_coeff - old[local_i]) > 1e-12:
                old[local_i] = new_coeff
                changed_vars.append(var)
                changed_coeffs.append(new_coeff)

        if changed_vars:
            for var, coeff in zip(changed_vars, changed_coeffs):
                var.Obj = coeff
            model.update()

    # =========================================================================
    # Gurobi progress callback
    # =========================================================================

    def _make_callback(self, s: int, phase_label: str):
        last_print = [0.0]
        lock       = self._progress_lock

        def cb(model, where):
            if where == GRB.Callback.MIPSOL:
                obj = model.cbGet(GRB.Callback.MIPSOL_OBJ)
                t   = model.cbGet(GRB.Callback.RUNTIME)
                with lock:
                    print(f"  [{phase_label} s={s:2d}] new incumbent {obj:,.0f}  ({t:.1f}s)",
                          flush=True)
            elif where == GRB.Callback.MIP:
                t = model.cbGet(GRB.Callback.RUNTIME)
                if t - last_print[0] >= 30:
                    last_print[0] = t
                    obj  = model.cbGet(GRB.Callback.MIP_OBJBST)
                    bnd  = model.cbGet(GRB.Callback.MIP_OBJBND)
                    nods = model.cbGet(GRB.Callback.MIP_NODCNT)
                    gap  = abs(obj - bnd) / max(abs(obj), 1e-10) * 100
                    with lock:
                        print(f"  [{phase_label} s={s:2d}] {t:.0f}s | obj {obj:,.0f} | "
                              f"bnd {bnd:,.0f} | gap {gap:.2f}% | nodes {nods:.0f}", flush=True)
        return cb

    # =========================================================================
    # Deviation helper
    # =========================================================================

    @staticmethod
    def _devs(sol_mat, x_bar_vec, is_binary, var_scale, participating_mask=None):
        diff = np.abs(sol_mat - x_bar_vec[np.newaxis, :])
        if participating_mask is not None:
            diff = diff.copy()
            diff[~participating_mask] = 0.0
        bin_diff      = diff[:, is_binary]
        cont_diff_rel = (diff / var_scale[np.newaxis, :])[:, ~is_binary]
        max_dev_bin   = bin_diff.max()       if bin_diff.size      > 0 else 0.0
        avg_dev_bin   = bin_diff.mean()      if bin_diff.size      > 0 else 0.0
        max_dev_cont  = cont_diff_rel.max()  if cont_diff_rel.size > 0 else 0.0
        avg_dev_cont  = cont_diff_rel.mean() if cont_diff_rel.size > 0 else 0.0
        return max_dev_bin, avg_dev_bin, max_dev_cont, avg_dev_cont

    # =========================================================================
    # solve() — initial solve + PH loop + post-processing
    # =========================================================================

    def solve(self):
        """Run initial solve → PH main loop → post-processing."""
        n_scenarios     = self.n_scenarios
        V               = self.V
        K               = self.K
        probs           = self.probs
        is_binary       = self.is_binary
        n_binary_na     = self.n_binary_na
        executor        = ThreadPoolExecutor(max_workers=self._n_workers_actual)

        # ------------------------------------------------------------------
        # Step 0: Initial solve (original objectives, no penalty)
        # ------------------------------------------------------------------
        total_ph_start = time()

        threads_per_sub = max(1, (os.cpu_count() or 4) // self._n_workers_actual)

        def _init_solve(s):
            model = self.scenarios[s]
            model.Params.MIPGap     = self.mip_gap
            model.Params.OutputFlag = 0
            model.Params.Threads    = threads_per_sub
            model.optimize()
            row = np.zeros(V, dtype=np.float64)
            if model.SolCount > 0:
                vidx   = self.scenario_vidx[s]
                vars_s = self.var_cache_flat[s]
                for _, (vi, var) in enumerate(zip(vidx, vars_s)):
                    if var is not None:
                        row[vi] = var.X
            obj_val = self.original_objectives[s].getValue() if model.SolCount > 0 else 0.0
            return s, row, obj_val, model.SolCount > 0

        futures_0   = {executor.submit(_init_solve, s): s for s in range(n_scenarios)}
        init_results_dict = {}
        with tqdm(total=n_scenarios, desc="Step 0 (initial solve)", unit="sc") as bar:
            for fut in as_completed(futures_0):
                s, row, obj_val, ok = fut.result()
                init_results_dict[s] = (row, obj_val)
                bar.update(1)
                if not ok:
                    tqdm.write(f"  [Step0 s={s}] WARNING: no solution found")

        init_obj = 0.0
        for s in range(n_scenarios):
            row, obj_val = init_results_dict[s]
            self.sol_matrix[s] = row
            init_obj += probs[s] * obj_val
        tqdm.write(f"Step 0 done — E[obj]: {init_obj:,.0f}")


        # ------------------------------------------------------------------
        # Set variable scales (from initial solutions)
        # ------------------------------------------------------------------
        for v in range(V):
            if not is_binary[v]:
                mx = np.abs(self.sol_matrix[:, v]).max()
                self.var_scale[v] = max(mx, 1000.0)

        # ------------------------------------------------------------------
        # Auto-calibrate rho (node-aware x̄ from the start)
        # ------------------------------------------------------------------
        x_bar      = self._compute_xbar()
        self.x_bar = x_bar
        diff       = self.sol_matrix - x_bar[np.newaxis, :]
        diff[~self._participating] = 0.0  # exclude non-participating (s, vi) pairs
        diff_sq    = diff ** 2
        bin_sq     = diff_sq[:, is_binary]
        cont_sq    = (diff / self.var_scale[np.newaxis, :]) ** 2
        cont_sq    = cont_sq[:, ~is_binary]

        sum_sq_bin  = bin_sq.sum()
        sum_sq_cont = cont_sq.sum()
        if sum_sq_bin > 1e-10:
            self.rho_bin  = max(2.0 * self.penalty_fraction * abs(init_obj) / (sum_sq_bin  / n_scenarios), 10.0)
            self.rho_bin  = min(self.rho_bin, 0.05 * abs(init_obj))  # cap at 5% of obj to avoid numerical issues
        if sum_sq_cont > 1e-10:
            self.rho_cont = max(2.0 * self.penalty_fraction * abs(init_obj) / (sum_sq_cont / n_scenarios), 1.0)
        self.rho_max_bin  = self.rho_max_obj_fraction * abs(init_obj)
        self.rho_max_cont = self.rho_max_obj_fraction * abs(init_obj)
        self._rho_bin_calibrated  = self.rho_bin
        self._rho_cont_calibrated = self.rho_cont
        print(f"Auto-calibrated rho_bin={self.rho_bin:.4g}, rho_cont={self.rho_cont:.4g}, "
              f"rho_max_bin={self.rho_max_bin:.4g}")

        # ------------------------------------------------------------------
        # Build augmented objectives (quadratic part uses calibrated rho)
        # ------------------------------------------------------------------
        self._build_augmented_objectives()

        rho_eff_vec = self._rho_eff_cont()
        for s in range(n_scenarios):
            self._update_obj_linear(s, x_bar, self.multipliers, self.rho_bin, rho_eff_vec)

        # ------------------------------------------------------------------
        # History initialisation
        # ------------------------------------------------------------------
        max_dev_bin0, avg_dev_bin0, max_dev_cont0, avg_dev_cont0 = self._devs(
            self.sol_matrix, x_bar, is_binary, self.var_scale, self._participating)
        max_dev0 = max(max_dev_bin0, max_dev_cont0)

        self.convergence_history  = [max_dev0]
        self.objective_history    = [init_obj]
        self.n_fixed_history      = [0]
        self.avg_dev_bin_history  = [avg_dev_bin0]
        self.avg_dev_cont_history = [avg_dev_cont0]
        self.disagree_history     = []

        print(f"\nStep 0 | Obj: {init_obj:12.2f} | MaxDev: {max_dev0:.2e} | "
              f"AvgBin: {avg_dev_bin0:.4f} AvgCont: {avg_dev_cont0:.4f} | Fixed: 0")

        # ------------------------------------------------------------------
        # PH main loop
        # ------------------------------------------------------------------
        k                  = 0
        converged          = False
        max_dev_bin        = max_dev_bin0
        max_dev_cont       = max_dev_cont0
        avg_dev_cont       = avg_dev_cont0
        _slam_triggered    = False
        _slam_triggered_at = None
        avg_dev_bin = avg_dev_bin0

        print(f"Starting PH iterations (max {K})")

        while k < K and not converged:
            iter_start = time()
            k += 1
            self._progress_counter[0] = 0
            phase_t = iter_start

            # 10a: Update linear coefficients (only changed entries)
            rho_eff_vec = self._rho_eff_cont()
            for s in range(n_scenarios):
                self._update_obj_linear(s, x_bar, self.multipliers, self.rho_bin, rho_eff_vec)

            # 10b: Solve subproblems in parallel
            def _ph_solve(s, _k=k):
                model = self.scenarios[s]
                model.Params.MIPGap     = self.mip_gap
                model.Params.OutputFlag = 0
                model.Params.Threads    = threads_per_sub
                model.Params.MIPFocus   = 1   # find feasible solutions fast; PH doesn't need optimality proof
                model.Params.TimeLimit = GRB.INFINITY
                vidx   = self.scenario_vidx[s]
                vars_s = self.var_cache_flat[s]
                for _, (vi, var) in enumerate(zip(vidx, vars_s)):
                    if var is not None:
                        var.Start = self.sol_matrix[s, vi]

                _t0       = time()
                _last_log = [_t0]
                _last_inc = [None]

                def _cb(__, where):
                    now = time()
                    elapsed = now - _t0
                    # Report new incumbents immediately (any time)
                    if where == GRB.Callback.MIPSOL and self.show_subproblem_incumbents:
                        try:
                            obj = model.cbGet(GRB.Callback.MIPSOL_OBJ)
                            if _last_inc[0] is None or abs(obj - _last_inc[0]) > 1e-6:
                                _last_inc[0] = obj
                                print(
                                    f"  [Iter {_k} s={s:02d} {elapsed:.1f}s] "
                                    f"new incumbent  obj={obj:.4e}",
                                    flush=True,
                                )
                        except Exception:
                            pass
                        return
                    # Report B&B progress every 2 s for slow solves (> 1 s)
                    if where != GRB.Callback.MIP or not self.show_subproblem_mip_progress:
                        return
                    if elapsed < 1.0 or now - _last_log[0] < 2.0:
                        return
                    _last_log[0] = now
                    try:
                        best = model.cbGet(GRB.Callback.MIP_OBJBST)
                        bnd  = model.cbGet(GRB.Callback.MIP_OBJBND)
                        gap  = abs(bnd - best) / (abs(best) + 1e-10)
                        node = int(model.cbGet(GRB.Callback.MIP_NODCNT))
                        print(
                            f"  [Iter {_k} s={s:02d} {elapsed:.0f}s] "
                            f"gap={gap:.1%}  best={best:.4e}  bnd={bnd:.4e}  nodes={node}",
                            flush=True,
                        )
                    except Exception:
                        pass

                model.optimize(_cb)

                rescued = False
                if model.SolCount == 0:
                    aug_obj = model.getObjective()
                    model.setObjective(self.original_objectives[s], GRB.MAXIMIZE)
                    model.Params.MIPGap = max(self.mip_gap, 0.05)
                    model.optimize()
                    rescued = model.SolCount > 0
                    model.setObjective(aug_obj, GRB.MAXIMIZE)

                row     = self.sol_matrix[s].copy()
                obj_val = 0.0
                if model.SolCount > 0:
                    for _, (vi, var) in enumerate(zip(vidx, vars_s)):
                        if var is not None:
                            row[vi] = self.fix_val[vi] if self.fixed[vi] else var.X
                    obj_val = self.original_objectives[s].getValue()
                elapsed_total = time() - _t0
                return s, row, obj_val, (model.SolCount > 0), rescued, elapsed_total

            futures_ph = {executor.submit(_ph_solve, s): s for s in range(n_scenarios)}
            fut_start = {f: time() for f in futures_ph}
            ph_results_dict = {}
            pending = set(futures_ph.keys())
            last_hb = time()
            hb_every_sec = 5.0
            last_progress = 0
            last_completion = time()
            last_stall_alert = time()
            stall_alert_every_sec = 15.0

            while pending:
                done, pending = wait(pending, timeout=1.0, return_when=FIRST_COMPLETED)

                for fut in done:
                    s, row, obj_val, has_sol, rescued, elapsed_total = fut.result()
                    ph_results_dict[s] = (s, row, obj_val, has_sol)
                    last_completion = time()
                    if not has_sol:
                        print(f"  [Iter {k} s={s}] NO SOL (status={self.scenarios[s].Status})", flush=True)
                    elif rescued:
                        print(f"  [Iter {k} s={s}] rescued", flush=True)
                    if elapsed_total >= 10.0:
                        print(f"  [Iter {k} s={s}] solved in {elapsed_total:.1f}s", flush=True)

                done_count = len(ph_results_dict)
                if done_count != last_progress and (done_count % 5 == 0 or done_count == n_scenarios):
                    print(f"  [Iter {k}] progress: {done_count}/{n_scenarios} done, {len(pending)} running", flush=True)
                    last_progress = done_count

                now = time()
                if pending and (now - last_hb) >= hb_every_sec:
                    pending_ids = sorted(futures_ph[f] for f in pending)
                    pending_preview = ",".join(f"{pid:02d}" for pid in pending_ids[:12])
                    if len(pending_ids) > 12:
                        pending_preview += ",..."
                    oldest_wait = max(now - fut_start[f] for f in pending)
                    print(
                        f"  [Iter {k}] heartbeat: {done_count}/{n_scenarios} done, "
                        f"{len(pending_ids)} running, oldest={oldest_wait:.1f}s | pending: {pending_preview}",
                        flush=True,
                    )
                    last_hb = now

                if pending and (now - last_completion) >= stall_alert_every_sec and (now - last_stall_alert) >= stall_alert_every_sec:
                    pending_with_elapsed = sorted(
                        ((futures_ph[f], now - fut_start[f]) for f in pending),
                        key=lambda x: x[1],
                        reverse=True,
                    )
                    top_slow = ", ".join(f"s={sid:02d}:{elap:.1f}s" for sid, elap in pending_with_elapsed[:5])
                    print(
                        f"  [Iter {k}] no completions for {now - last_completion:.1f}s; "
                        f"{len(pending)} still running | slowest: {top_slow}",
                        flush=True,
                    )
                    last_stall_alert = now

            print(f"  [Iter {k}] subproblem solves complete in {time() - phase_t:.1f}s", flush=True)
            phase_t = time()
            ph_results = [ph_results_dict[s] for s in range(n_scenarios)]

            new_sol_matrix = np.empty_like(self.sol_matrix)
            iter_obj       = 0.0
            n_no_sol_iter = 0
            for s, row, obj_val, has_solution in ph_results:
                new_sol_matrix[s] = row
                iter_obj          += probs[s] * obj_val
                if not has_solution:
                    n_no_sol_iter += 1
            self.sol_matrix = new_sol_matrix
            self.objective_history.append(iter_obj)

            print(f"  [Iter {k}] result aggregation complete in {time() - phase_t:.1f}s", flush=True)
            phase_t = time()

            if n_no_sol_iter > 0:
                print(f"  WARNING: {n_no_sol_iter}/{n_scenarios} subproblems had no incumbent in iter {k}.")

            # 10c: Update x̄ — node-aware (single vectorised operation)
            x_bar      = self._compute_xbar()
            self.x_bar = x_bar

            print(f"  [Iter {k}] x_bar update complete in {time() - phase_t:.1f}s", flush=True)
            phase_t = time()

            # 10d: Variable fixing — node-aware consensus check
            # For a node variable, we check agreement only within the scenarios
            # that share that node (not across all 81).
            newly_fixed = 0
            if n_binary_na > 0:
                bin_cols = np.where(is_binary & ~self.fixed)[0]
                fix_scan_start = time()
                fix_scan_last  = fix_scan_start
                n_bin_cols     = len(bin_cols)
                if n_bin_cols > 0:
                    print(f"  [Iter {k}] variable fixing scan start ({n_bin_cols} binary NA vars)", flush=True)
                for i, vi in enumerate(bin_cols, start=1):
                    node  = self._vi_to_node[vi]
                    scens = self._node_scens[node]
                    vals  = self.sol_matrix[scens, vi]
                    ref_v = np.round(vals[0])
                    if np.all(np.abs(vals - ref_v) < self.epsilon_bin):
                        self.unanimous_count[vi] += 1
                    else:
                        self.unanimous_count[vi] = 0
                    now_fix = time()
                    if now_fix - fix_scan_last >= 5.0:
                        print(
                            f"  [Iter {k}] variable fixing scan: {i}/{n_bin_cols} checked "
                            f"({now_fix - fix_scan_start:.1f}s)",
                            flush=True,
                        )
                        fix_scan_last = now_fix
                to_fix = bin_cols[self.unanimous_count[bin_cols] >= self.fix_threshold]
                n_to_fix = len(to_fix)
                if n_to_fix > 0:
                    print(f"  [Iter {k}] applying bounds for {n_to_fix} newly fixed vars", flush=True)
                apply_fix_last = time()
                for i, vi in enumerate(to_fix, start=1):
                    node  = self._vi_to_node[vi]
                    scens = self._node_scens[node]
                    fv    = float(np.round(self.sol_matrix[scens[0], vi]))
                    self.fixed[vi]   = True
                    self.fix_val[vi] = fv
                    newly_fixed     += 1
                    for s, local_i in self._vi_occurrences[vi]:
                        v_obj = self.var_cache_flat[s][local_i]
                        if v_obj is not None:
                            v_obj.LB = fv
                            v_obj.UB = fv
                    now_apply = time()
                    if now_apply - apply_fix_last >= 5.0:
                        print(
                            f"  [Iter {k}] applying fixed bounds: {i}/{n_to_fix} "
                            f"({now_apply - fix_scan_start:.1f}s)",
                            flush=True,
                        )
                        apply_fix_last = now_apply
                if n_to_fix > 0:
                    self.multipliers[:, to_fix] = 0.0

            print(f"  [Iter {k}] variable fixing complete in {time() - phase_t:.1f}s", flush=True)
            phase_t = time()

            # 10e: Adaptive slam trigger
            if not _slam_triggered and k >= self.min_iters_before_slam:
                recent_disagree = self.disagree_history[-self.slam_stall_window:]
                recent_devs     = self.convergence_history[-self.slam_stall_window:]
                peak_disagree   = max(self.disagree_history) if self.disagree_history else 1
                if len(recent_disagree) >= self.slam_stall_window:
                    disagree_stagnated  = min(recent_disagree) >= (1 - self.slam_stall_tol) * recent_disagree[0]
                    dev_stagnated       = (recent_devs[0] <= 0 or
                                          (recent_devs[0] - min(recent_devs)) / recent_devs[0] < self.slam_stall_tol)
                    made_progress       = recent_disagree[-1] < self.slam_max_dev_threshold * peak_disagree
                    if disagree_stagnated and dev_stagnated and made_progress:
                        _slam_triggered    = True
                        _slam_triggered_at = k
                        tqdm.write(f"  [Iter {k}] Slam triggered — "
                                   f"Disagree stagnated ({recent_disagree[0]}→{min(recent_disagree)}, "
                                   f"peak={peak_disagree}) and MaxDev stagnated "
                                   f"({recent_devs[0]:.3e}→{min(recent_devs):.3e})")

            # 10e: Cycle-breaking slam
            n_slammed = 0
            if _slam_triggered:
                slam_hist_last = time()
                n_hist = len(self.xbar_binary_history)
                for i, (vi, hist) in enumerate(self.xbar_binary_history.items(), start=1):
                    if self.fixed[vi] or len(hist) < self.slam_stall_window:
                        continue
                    recent = hist[-self.slam_stall_window:]
                    rmin, rmax = min(recent), max(recent)
                    rmean      = float(np.mean(recent))
                    # (a) Oscillating var with wide swings — slam to rounded mean.
                    # (b) Stuck-flat var at a non-integral consensus value (e.g.
                    #     tied 0.5 across scenarios) — also slam to rounded mean
                    #     so PH can move on instead of spinning.
                    is_oscillating = (rmax - rmin) > 0.15
                    is_stuck_flat  = ((rmax - rmin) < 0.05
                                     and min(rmean, 1.0 - rmean) > self.epsilon_bin)
                    if is_oscillating or is_stuck_flat:
                        slam_v = float(np.round(rmean))
                        self.fixed[vi]   = True
                        self.fix_val[vi] = slam_v
                        n_slammed       += 1
                    now_slam = time()
                    if now_slam - slam_hist_last >= 5.0:
                        print(f"  [Iter {k}] slam scan: {i}/{n_hist} checked", flush=True)
                        slam_hist_last = now_slam

            # Consistency repair on newly slammed variables: z=1 must have ≥1 h=1
            _to_apply: set[int] = set()
            if n_slammed > 0:
                _z_slam: dict[tuple, int] = {}
                _h_slam: dict[tuple, list] = {}
                slam_repair_last = time()
                n_all_na = len(self.all_na_names_ordered)
                for i, (vi2, qname2) in enumerate(zip(range(n_all_na), self.all_na_names_ordered), start=1):
                    if not (is_binary[vi2] and self.fixed[vi2]):
                        continue
                    vname2 = qname2.split("@")[0]  # strip @node
                    br2    = vname2.index("[")
                    pfx2   = vname2[:br2]
                    inner2 = vname2[br2+1:-1]
                    if pfx2 == "z":
                        u2, ts2 = inner2.rsplit(",", 1)
                        _z_slam[(u2, int(ts2))] = vi2
                    elif pfx2 == "h":
                        parts2 = inner2.rsplit(",", 2)
                        key2 = (parts2[0], int(parts2[1]))
                        _h_slam.setdefault(key2, []).append((vi2, int(parts2[2])))
                    now_repair = time()
                    if now_repair - slam_repair_last >= 5.0:
                        print(f"  [Iter {k}] slam repair scan: {i}/{n_all_na} vars", flush=True)
                        slam_repair_last = now_repair
                for (u2, ts2), z_vi2 in _z_slam.items():
                    if self.fix_val[z_vi2] < 0.5:
                        _to_apply.add(z_vi2)
                        continue
                    hlist2 = _h_slam.get((u2, ts2), [])
                    if any(self.fix_val[hv2] >= 0.5 for hv2, _ in hlist2):
                        _to_apply.add(z_vi2)
                        for hv2, _ in hlist2:
                            _to_apply.add(hv2)
                        continue
                    if hlist2:
                        best_hvi2, _ = max(hlist2, key=lambda hv2: self.x_bar[hv2[0]])
                        if self.x_bar[best_hvi2] > 0.0:
                            self.fix_val[best_hvi2] = 1.0
                            _to_apply.add(best_hvi2)
                        else:
                            self.fix_val[z_vi2] = 0.0
                    else:
                        self.fix_val[z_vi2] = 0.0
                    _to_apply.add(z_vi2)
                    for hv2, _ in hlist2:
                        _to_apply.add(hv2)

            # Apply bounds for newly slammed + repaired variables
            apply_slam_last = time()
            n_apply_slam = len(_to_apply)
            for i, vi2 in enumerate(_to_apply, start=1):
                fv2 = self.fix_val[vi2]
                for s, local_i in self._vi_occurrences[vi2]:
                    v_obj = self.var_cache_flat[s][local_i]
                    if v_obj is not None:
                        v_obj.LB = fv2
                        v_obj.UB = fv2
                now_apply_slam = time()
                if now_apply_slam - apply_slam_last >= 5.0:
                    print(
                        f"  [Iter {k}] applying slam bounds: {i}/{n_apply_slam}",
                        flush=True,
                    )
                    apply_slam_last = now_apply_slam
            if n_apply_slam > 0:
                apply_slam_idx = np.fromiter(_to_apply, dtype=np.int32, count=n_apply_slam)
                self.multipliers[:, apply_slam_idx] = 0.0

            # Update binary x_bar history (node-specific x̄)
            for vi in list(self.xbar_binary_history.keys()):
                if not self.fixed[vi]:
                    self.xbar_binary_history[vi].append(float(x_bar[vi]))

            self.n_fixed_history.append(int(self.fixed.sum()))

            # 10f: Convergence check — apply participating mask before computing deviations
            unfixed_mask = ~self.fixed
            diff_uf      = np.abs(self.sol_matrix - x_bar[np.newaxis, :])
            diff_uf[~self._participating] = 0.0  # zero out non-participating (s, vi)
            diff_uf[:, self.fixed]        = 0.0
            bin_mask_uf  = is_binary & unfixed_mask
            cont_mask_uf = ~is_binary & unfixed_mask

            max_dev_bin  = diff_uf[:, bin_mask_uf].max()  if bin_mask_uf.any()  else 0.0
            avg_dev_bin  = diff_uf[:, bin_mask_uf].mean() if bin_mask_uf.any()  else 0.0
            rel_cont     = diff_uf[:, cont_mask_uf] / self.var_scale[np.newaxis, cont_mask_uf]
            max_dev_cont = rel_cont.max()  if cont_mask_uf.any() else 0.0
            avg_dev_cont = rel_cont.mean() if cont_mask_uf.any() else 0.0

            self.avg_dev_bin_history.append(avg_dev_bin)
            self.avg_dev_cont_history.append(avg_dev_cont)
            max_dev = max(max_dev_bin, max_dev_cont)
            self.convergence_history.append(max_dev)

            bin_disagree = int((diff_uf[:, bin_mask_uf] > self.epsilon_bin).sum()) if bin_mask_uf.any() else 0
            self.disagree_history.append(bin_disagree)
            converged    = (max_dev_bin <= self.epsilon_bin) and (max_dev_cont <= self.epsilon_cont)

            # Tail-stagnation escape: if disagreement is small but flat for many
            # iterations, force-fix the remaining binary NA vars to rounded x_bar.
            # This prevents very long tails where a handful of binaries never settle.
            tail_fixed = 0
            if (
                not converged
                and _slam_triggered
                and bin_disagree > 0
                and bin_disagree <= self.tail_fix_max_disagree
                and len(self.disagree_history) >= self.tail_fix_window
            ):
                recent_disagree = self.disagree_history[-self.tail_fix_window:]
                if min(recent_disagree) == max(recent_disagree):
                    rem_bin = np.where(is_binary & ~self.fixed)[0]
                    if rem_bin.size > 0:
                        print(
                            f"  [Iter {k}] tail stall detected: disagree={bin_disagree} "
                            f"for {self.tail_fix_window} iters -> force-fixing {rem_bin.size} vars",
                            flush=True,
                        )
                    for vi in rem_bin:
                        fv = float(np.round(x_bar[vi]))
                        self.fixed[vi] = True
                        self.fix_val[vi] = fv
                        tail_fixed += 1
                        for s, local_i in self._vi_occurrences[vi]:
                            v_obj = self.var_cache_flat[s][local_i]
                            if v_obj is not None:
                                v_obj.LB = fv
                                v_obj.UB = fv
                    if tail_fixed > 0:
                        self.multipliers[:, rem_bin] = 0.0
                        newly_fixed += tail_fixed
                        converged = True

            print(f"  [Iter {k}] convergence metrics complete in {time() - phase_t:.1f}s", flush=True)
            phase_t = time()

            elapsed_iter = time() - iter_start
            slam_state = (
                f"@{_slam_triggered_at}" if _slam_triggered
                else (f"rdy<{self.slam_max_dev_threshold}"
                      if max_dev_bin < self.slam_max_dev_threshold else "wait")
            )
            print(
                f"Iter {k:4d}/{K} | obj={iter_obj:.3e} | BinDev={max_dev_bin:.2e} | "
                f"Disagree={bin_disagree} | Fixed={self.fixed.sum()}(+{newly_fixed}f+{n_slammed}s) | "
                f"rho_bin={self.rho_bin:.1e} | slam={slam_state} | t={elapsed_iter:.1f}s",
                flush=True,
            )

            # 10g: Multiplier update — zero out non-participating (s, vi) pairs
            rho_vec = np.where(is_binary, self.rho_bin, self.rho_cont / (self.var_scale ** 2))
            self.multipliers += rho_vec[np.newaxis, :] * (self.sol_matrix - x_bar[np.newaxis, :])
            self.multipliers[~self._participating] = 0.0  # only penalise participating pairs
            self.multipliers[:, self.fixed]        = 0.0

            print(f"  [Iter {k}] multiplier update complete in {time() - phase_t:.1f}s", flush=True)
            phase_t = time()

            # 10h: Adaptive rho increase if stalled
            if k >= self.stall_window and len(self.disagree_history) >= self.stall_window:
                recent_d = self.disagree_history[-self.stall_window:]
                if min(recent_d) > 0.8 * recent_d[0]:
                    self.rho_bin  = min(self.rho_bin  * self.rho_increase, self.rho_max_bin)
                    self.rho_cont = min(self.rho_cont * self.rho_increase, self.rho_max_cont)
                elif recent_d[-1] < 0.8 * recent_d[0]:
                    self.rho_bin  = max(self.rho_bin  / self.rho_increase, self._rho_bin_calibrated)
                    self.rho_cont = max(self.rho_cont / self.rho_increase, self._rho_cont_calibrated)

            print(f"  [Iter {k}] rho adaptation complete in {time() - phase_t:.1f}s", flush=True)

        self.converged    = converged
        self.n_iters      = k
        self.total_time   = time() - total_ph_start
        self.max_dev_bin  = max_dev_bin
        self.avg_dev_bin  = avg_dev_bin
        self.max_dev_cont = max_dev_cont
        self.avg_dev_cont = avg_dev_cont

        self._post_process(executor)
        executor.shutdown(wait=False)

    # =========================================================================
    # _post_process — robust q LP + per-scenario re-solve
    # =========================================================================

    def _post_process(self, executor):
        """Robust q LP across all scenarios, then per-scenario re-solve for stage-3 decisions."""
        n_scenarios = self.n_scenarios
        V           = self.V
        T           = self.T
        probs       = self.probs
        is_binary   = self.is_binary

        print("\n" + "=" * 70)
        print("POST-PROCESSING: robust q LP (all scenarios) + per-scenario re-solve")
        print("=" * 70)
        eval_start = time()

        # --- Binary consensus ---
        bin_consensus: dict[int, float] = {}
        for vi in range(V):
            if not is_binary[vi]:
                continue
            bin_consensus[vi] = self.fix_val[vi] if self.fixed[vi] else float(np.round(self.x_bar[vi]))

        n_hard = int(self.fixed.sum())
        n_soft = len(bin_consensus) - n_hard
        print(f"Consensus binary NA vars: {n_hard} hard-fixed by PH, {n_soft} from rounded x̄")

        self.bin_consensus = bin_consensus

        # Consistency repair: h[u,ss,th]=1 is impossible if z[u,ss]=0 in the same node.
        # PH fixes z and h independently so they can end up contradicting each other.
        # Note: h may be at a finer node granularity than z (e.g. h is stage-3, z is stage-2),
        # so we derive z's node key from h's node key by stripping the extra stage component.
        n_repaired = 0
        for vi, qname in enumerate(self.all_na_names_ordered):
            if not is_binary[vi] or bin_consensus.get(vi, 0.0) < 0.5:
                continue
            vname = qname.split("@")[0]
            if not vname.startswith("h["):
                continue
            h_nk  = qname.split("@")[1]
            inner = vname[2:-1]                        # "u,ss,th"
            u_str, ss_str, _ = inner.rsplit(",", 2)
            ss    = int(ss_str)
            # Derive the node key at the stage ss belongs to (z's granularity)
            p = h_nk.split("_")                        # e.g. ["s3","normal","good","bad"]
            if ss < self.stage0_end:
                z_nk = "s0"
            elif ss < self.stage1_end:
                z_nk = f"s1_{p[1]}"
            elif ss < self.stage2_end:
                z_nk = f"s2_{p[1]}_{p[2]}"
            else:
                z_nk = h_nk                            # same stage — key matches directly
            z_vi = self.name_to_idx.get(f"z[{u_str},{ss_str}]@{z_nk}")
            if z_vi is not None and bin_consensus.get(z_vi, 0.0) < 0.5:
                bin_consensus[vi] = 0.0
                n_repaired += 1
        if n_repaired:
            print(f"  [post-process] consistency repair: zeroed {n_repaired} h vars where z=0")

        # Apply binary bounds + restore penalty-free objectives
        for s in range(n_scenarios):
            model  = self.scenarios[s]
            vidx   = self.scenario_vidx[s]
            vars_s = self.var_cache_flat[s]
            for _, (vi, var) in enumerate(zip(vidx, vars_s)):
                if var is None or not is_binary[vi]:
                    continue
                val = bin_consensus[vi]
                var.LB = val
                var.UB = val
            model.setObjective(self.original_objectives[s], GRB.MAXIMIZE)
            model.update()

        # Parse consensus into stocking / harvest event sets (strip @node)
        _stock_na  = {}
        _harv_na   = {}
        _hexist_na = {}
        for vi, qname in enumerate(self.all_na_names_ordered):
            if not is_binary[vi] or bin_consensus.get(vi, 0.0) < 0.5:
                continue
            vname = qname.split("@")[0]  # strip @node
            br    = vname.index("[")
            pfx   = vname[:br]
            inner = vname[br+1:-1]
            if pfx == "z":
                u, ts = inner.rsplit(",", 1)
                _stock_na[(u, int(ts))] = True
            elif pfx == "h":
                p = inner.rsplit(",", 2)
                _harv_na[(p[0], int(p[1]), int(p[2]))] = True
            elif pfx == "h_exist":
                u, ts = inner.rsplit(",", 1)
                _hexist_na[(u, int(ts))] = True

        self._stock_na  = _stock_na
        self._harv_na   = _harv_na
        self._hexist_na = _hexist_na

        print(f"Active NA cohorts: {len(_stock_na)} stocked, "
              f"{len(_harv_na)} NA harvests, {len(_hexist_na)} exist. harvests")

        # --- Alive flags ---
        _alive: dict = {}
        _ref   = self.milp_objects[0]
        for (u, ss) in _stock_na:
            harvest_times = [t for (uh, ssh, t) in self._harv_na if uh == u and ssh == ss]
            t_h = harvest_times[0] if harvest_times else T
            alive_periods = [t for t in range(ss, t_h) if t < T]
            if alive_periods:
                _alive[(u, ss)] = alive_periods

        _alive_exist: dict = {}
        for u in _ref.U_exist:
            harvest_times = [t for (uh, t) in self._hexist_na if uh == u]
            t_h = harvest_times[0] if harvest_times else T
            alive_periods = [t for t in range(0, t_h) if t < T]
            if alive_periods:
                _alive_exist[u] = alive_periods

        # ================================================================
        # Four-stage stochastic LP for q (deterministic equivalent with NACs)
        # ================================================================
        recourse_q = gb.Model("recourse_q")
        recourse_q.Params.OutputFlag = 0
        T_end = T - 1
        Q_MAX_ROB = 1_000_000
        MAB_PENALTY = 1000

        _q_keys: set = set()
        for (u, ss) in _stock_na:
            _q_keys.add((u, ss))

        # Helpers: derive which tree node scenario s is in at month ss.
        # Generalised — works for any number of branching stages, with or
        # without a deterministic prefix.
        def _node_key(s: int, ss: int) -> str:
            return self._node_key_for(s, ss)

        def _z_cons(s: int, u, ss: int) -> float:
            nk = _node_key(s, ss)
            vi = self.name_to_idx.get(f"z[{u},{ss}]@{nk}")
            return bin_consensus.get(vi, 0.0) if vi is not None else 0.0

        # Build the set of (node_key, u, ss) for all node-cohort combinations where z=1.
        # q_rob is indexed by node so that the NAC reads:
        #   all scenarios s passing through node n share q_s[u,ss] = q_rob[n, u, ss].
        _q_node_keys: set = set()
        for s in range(n_scenarios):
            for (u, ss) in _q_keys:
                if _z_cons(s, u, ss) >= 0.5:
                    _q_node_keys.add((_node_key(s, ss), u, ss))

        # UB per (node_key, u, ss): min over scenarios in that node.
        _q_node_ub: dict = {}
        for (nk, u, ss) in _q_node_keys:
            min_ub = Q_MAX_ROB
            for _ms in self.milp_objects:
                if (u, ss) in _ms.variables["q"]:
                    min_ub = min(min_ub, _ms.variables["q"][u, ss].UB)
            _q_node_ub[(nk, u, ss)] = min_ub

        _q_node_keys_list = list(_q_node_keys)
        q_rob = recourse_q.addVars(
            _q_node_keys_list,
            lb=0,
            ub=[_q_node_ub[k] for k in _q_node_keys_list],
            name=lambda k: f"q_rob_{k[0]}_{k[1]}_{k[2]}"
        )

        _ref0 = self.milp_objects[0]
        mab_slack_loc = recourse_q.addVars(
            _ref0.L_list, range(T), self.n_scenarios,
            lb=0, name="mab_slack_loc"
        )
        mab_slack_reg = recourse_q.addVars(
            range(T), self.n_scenarios,
            lb=0, name="mab_slack_reg"
        )
        dens_slack_new = recourse_q.addVars(
            _ref0.U, range(T), self.n_scenarios,
            lb=0, name="dens_slack_new"
        )
        dens_slack_exist = recourse_q.addVars(
            _ref0.U_exist, range(T), self.n_scenarios,
            lb=0, name="dens_slack_exist"
        )
        recourse_q.update()

        _obj = gb.LinExpr()

        for s in range(n_scenarios):
            _s_ref = self.milp_objects[s]
            p_s = self.probs[s]

            q_s = recourse_q.addVars(_s_ref.U, _s_ref.Tset, lb=0, name=f"q_{s}")

            for u, ss in _q_keys:
                nk = _node_key(s, ss)
                if _z_cons(s, u, ss) >= 0.5:
                    # NAC: scenarios sharing node nk must agree on q[u,ss]
                    recourse_q.addConstr(q_s[u, ss] == q_rob[nk, u, ss], name=f"q_link_{s}_{u}_{ss}")
                else:
                    # Cohort not stocked in this branch — fix q=0
                    recourse_q.addConstr(q_s[u, ss] == 0, name=f"q_zero_{s}_{u}_{ss}")

            _bpf_alive_s: dict = {}
            for (u, ss), alive_t in _alive.items():
                for t in alive_t:
                    if (u, t) not in _bpf_alive_s:
                        _bpf_alive_s[(u, t)] = []
                    _bpf_alive_s[(u, t)].append((ss, float(_s_ref.bpf[ss][t])))

            def _biom_new(u, t):
                return gb.quicksum(c * q_s[u, ss] for ss, c in _bpf_alive_s.get((u, t), []))

            def _biom_exist(u, t):
                if u in _alive_exist and t in _alive_exist[u]:
                    return float(_s_ref.b_exist[u][t])
                return 0.0

            for u in _s_ref.U:
                cap_kg = _s_ref.density_limit * _s_ref.vol[u]
                for t in range(T):
                    recourse_q.addConstr(
                        _biom_new(u, t) <= cap_kg + dens_slack_new[u, t, s],
                        name=f"dens_{s}_{u}_{t}"
                    )
            for u in _s_ref.U_exist:
                cap_kg = _s_ref.density_limit * _s_ref.vol[u]
                for t in range(T):
                    recourse_q.addConstr(
                        _biom_exist(u, t) <= cap_kg + dens_slack_exist[u, t, s],
                        name=f"dens_exist_{s}_{u}_{t}"
                    )

            for l in _s_ref.L_list:
                for t in range(T):
                    units_in_loc = [u for u in _s_ref.U if _s_ref.loc_of[u] == l]
                    exist_units_in_loc = [u for u in _s_ref.U_exist if _s_ref.loc_of[u] == l]
                    biomass_loc_t = (
                        gb.quicksum(_biom_new(u, t) for u in units_in_loc) +
                        sum(_biom_exist(u, t) for u in exist_units_in_loc)
                    )
                    recourse_q.addConstr(biomass_loc_t <= _s_ref.loc_mab[l] + mab_slack_loc[l, t, s],
                                         name=f"mab_loc_{s}_{l}_{t}")

            for t in range(T):
                biomass_reg_t = (
                    gb.quicksum(_biom_new(u, t) for u in _s_ref.U) +
                    sum(_biom_exist(u, t) for u in _s_ref.U_exist)
                )
                recourse_q.addConstr(biomass_reg_t <= self.regional_mab + mab_slack_reg[t, s],
                                     name=f"mab_reg_{s}_{t}")

            # end_bio_min constraint removed (see IP.py): terminal_val terms in
            # objective already incentivise keeping fish alive at T_end.

            def _price_new(ss, tt):
                return _s_ref._price_for_kg(_s_ref.wpf[ss][tt] / 1000.0)

            revenue = gb.quicksum(
                float(_s_ref.bpf[ss][t]) * q_s[u, ss] * _price_new(ss, t) * _s_ref.df[t]
                for (u, ss, t) in self._harv_na.keys()
                if (u, ss) in q_s
            )
            exist_harv_revenue = sum(
                float(_s_ref.b_exist[u][t_h])
                * _s_ref._price_for_kg(_s_ref.w_exist[u][t_h] / 1000.0)
                * _s_ref.df[t_h]
                for (u, t_h) in _hexist_na.keys()
                if t_h < T
            )
            smolt_cost = gb.quicksum(
                _s_ref.smolt_cost_per_head * _s_ref.df[ss] * q_s[u, ss]
                for u, ss in q_s.keys()
            )
            feed_cost = gb.quicksum(
                _s_ref.feed_cost_per_kg_month * _s_ref.df[t] * float(_s_ref.bpf[ss][t]) * q_s[u, ss]
                for (u, ss), alive_t in _alive.items()
                for t in alive_t
            ) + sum(
                _s_ref.feed_cost_per_kg_month * _s_ref.df[t] * float(_s_ref.b_exist[u][t])
                for u, alive_t in _alive_exist.items()
                for t in alive_t
            )
            terminal_value = gb.quicksum(
                _s_ref.terminal_value_per_kg * float(_s_ref.bpf[ss][T_end]) * q_s[u, ss]
                for (u, ss), alive_t in _alive.items()
                if T_end in alive_t
            ) + sum(
                _s_ref.terminal_value_per_kg * float(_s_ref.b_exist[u][T_end])
                for u, alive_t in _alive_exist.items()
                if T_end in alive_t
            )
            _obj.add(p_s * (revenue + exist_harv_revenue - smolt_cost - feed_cost + terminal_value))

            _obj.add(-p_s * MAB_PENALTY * gb.quicksum(
                mab_slack_loc[l, t, s] for l in _s_ref.L_list for t in range(T)))
            _obj.add(-p_s * MAB_PENALTY * gb.quicksum(
                mab_slack_reg[t, s] for t in range(T)))
            _obj.add(-p_s * MAB_PENALTY * gb.quicksum(
                dens_slack_new[u, t, s] for u in _s_ref.U for t in range(T)))
            _obj.add(-p_s * MAB_PENALTY * gb.quicksum(
                dens_slack_exist[u, t, s] for u in _s_ref.U_exist for t in range(T)))
            # slack_end_bio_s removed with end_biomass_min constraint

        recourse_q.setObjective(_obj, GRB.MAXIMIZE)
        recourse_q.optimize()

        _rob_lp_time = time() - eval_start
        rob_feasible = recourse_q.SolCount > 0
        if not rob_feasible:
            print("WARNING: Two-stage recourse LP infeasible.")
            recourse_q.computeIIS()
            recourse_q.write("recourse_iis.ilp")
            q_robust_vals = {}
        else:
            # q_robust_vals keyed by (node_key, u, ss)
            q_robust_vals = {k: v.X for k, v in q_rob.items()}
            self.q_robust_vals = q_robust_vals
            print(f"Four-stage stochastic LP in extensive form solved in {_rob_lp_time:.2f}s. Robust q found.")

        # --- Fix robust q in each scenario, re-solve for stage-3 decisions ---
        def _eval_solve(s):
            model = self.scenarios[s]
            milp  = self.milp_objects[s]
            start = time()

            q_vars = milp.variables["q"]
            # q_robust_vals is keyed by (node_key, u, ss); look up the node for this scenario.
            for (nk, u, ss), q_val in q_robust_vals.items():
                if _node_key(s, ss) != nk:
                    continue  # this entry belongs to a different node
                if (u, ss) not in q_vars:
                    continue
                if _z_cons(s, u, ss) < 0.5:
                    q_val = 0.0  # cohort not stocked in this branch
                elif q_val < 0.5:
                    q_val = 0.0
                elif q_val < 1.0:
                    q_val = 1.0
                q_vars[u, ss].lb = q_val
                q_vars[u, ss].ub = q_val

            model.optimize()
            elapsed  = time() - start
            feasible = model.SolCount > 0
            obj_val  = model.ObjVal if feasible else 0.0

            if not feasible:
                # Compute IIS to identify conflicting fixed bounds.
                model.computeIIS()
                with self._progress_lock:
                    if not getattr(self, '_iis_written', False):
                        self._iis_written = True
                        iis_file = f"iis_scenario_{s}.ilp"
                        model.write(iis_file)
                        tqdm.write(f"  [post-proc s={s}] INFEASIBLE — IIS written to {iis_file}")

                # Backtrack: iteratively relax IIS-flagged bounds until feasible.
                # A single pass may not resolve all conflicts because Gurobi's
                # IIS is one *minimal* infeasible subsystem, not all of them.
                MAX_RELAX_PASSES = 5
                total_relaxed = 0
                relax_pass    = 0
                while relax_pass < MAX_RELAX_PASSES:
                    n_relaxed = 0
                    for v in model.getVars():
                        if v.IISLB or v.IISUB:
                            v.LB = 0.0
                            v.UB = 1.0
                            n_relaxed += 1
                    if n_relaxed == 0:
                        break  # nothing flagged — shouldn't happen, be defensive
                    total_relaxed += n_relaxed
                    relax_pass    += 1
                    model.reset()
                    model.optimize()
                    feasible = model.SolCount > 0
                    if feasible:
                        break
                    # Still infeasible — find the next IIS and keep going.
                    model.computeIIS()

                elapsed = time() - start
                obj_val = model.ObjVal if feasible else 0.0
                if feasible:
                    tqdm.write(f"  [post-proc s={s}] feasible after {relax_pass} pass(es), "
                               f"{total_relaxed} bound(s) relaxed  obj={obj_val:,.0f}")
                else:
                    tqdm.write(f"  [post-proc s={s}] still infeasible after {MAX_RELAX_PASSES} "
                               f"passes ({total_relaxed} bound(s) relaxed) — dropping from eval")

            return s, feasible, obj_val, elapsed

        eval_results = list(executor.map(_eval_solve, range(n_scenarios)))

        n_feasible, n_infeasible = 0, 0
        eval_obj             = 0.0
        infeasible_scenarios = []

        for s, feasible, obj_val, elapsed in eval_results:
            if feasible:
                n_feasible += 1
                eval_obj   += probs[s] * obj_val
                status_str  = f"obj={obj_val:>14,.0f}"
            else:
                n_infeasible += 1
                infeasible_scenarios.append(s)
                status_str   = "INFEASIBLE"
            print(f"  s={s:2d}  {self.scenario_names[s][:40]:<40}  {status_str}  ({elapsed:.1f}s)")

        eval_time = time() - eval_start

        self.eval_obj             = eval_obj
        self.n_feasible           = n_feasible
        self.n_infeasible         = n_infeasible
        self.infeasible_scenarios = infeasible_scenarios

        print("\n" + "-" * 70)
        print(f"Feasibility : {n_feasible}/{n_scenarios} feasible"
              + (f"  [INFEASIBLE: {infeasible_scenarios}]" if infeasible_scenarios else "  [ALL OK]"))
        print(f"E[obj] robust q + re-solved: {eval_obj:,.2f}")
        print(f"E[obj] PH final iter:        {self.objective_history[-1]:,.2f}")
        print(f"Post-processing time: {eval_time:.1f}s")
        print("=" * 70)

    # =========================================================================
    # print_results()
    # =========================================================================

    def print_results(self):
        print("\n" + "=" * 70)
        print(f"{'CONVERGED' if self.converged else 'MAX ITERATIONS'} after {self.n_iters} iters ({self.total_time:.1f}s)")
        print(f"E[obj] post-processing:         {self.eval_obj:,.2f}  "
              f"({self.n_feasible}/{self.n_scenarios} feasible scenarios)")
        print(f"E[obj] PH final iter (aux):     {self.objective_history[-1]:,.2f}")
        print(f"Final max binary deviation:     {self.max_dev_bin:.2e}  (avg: {self.avg_dev_bin:.4f})")
        print(f"Variables fixed: {self.fixed.sum()} / {self.n_binary_na} binary NA vars")
        print(f"  (V={self.V} total: {self.n_binary_na} binary, {self.n_cont_na} continuous)")

        # --- Branch differentiation diagnostic ---
        # For each stage-1 and stage-2 variable that has a counterpart in all three branches,
        # report how many differ between bad/normal/good.
        print("\n--- Branch differentiation (stage-1 and stage-2 binary consensus) ---")
        # Build map: base_varname -> {node: consensus_val}
        branch_map: dict[str, dict[str, float]] = {}
        for vi, qn in enumerate(self.all_na_names_ordered):
            if not self.is_binary[vi]:
                continue
            vn, node = qn.split("@", 1)
            if node == "s0":
                continue  # stage-0 is shared by design
            branch_map.setdefault(vn, {})[node] = self.bin_consensus.get(vi, 0.0)

        # Count variables where bad != normal OR normal != good
        n_differ = 0
        examples = []
        for vn, node_vals in branch_map.items():
            vals = list(node_vals.values())
            if len(set(round(v) for v in vals)) > 1:
                n_differ += 1
                if len(examples) < 5:
                    examples.append(f"  {vn}: " + ", ".join(f"{nd}={v:.2f}" for nd, v in node_vals.items()))

        total_branch_vars = len(branch_map)
        print(f"  Branch-specific binary vars (stage-1/2): {total_branch_vars}")
        print(f"  Vars where branches differ in consensus:  {n_differ} / {total_branch_vars}")
        if n_differ == 0:
            print("  → All branches converged to the SAME binary decisions.")
            print("    This means the recourse model found one plan optimal across all temperature outcomes.")
            print("    Run more iterations or lower fix_threshold to allow more branch differentiation.")
        else:
            print(f"  → {n_differ} variables differ between branches (recourse IS adapting).")
            for ex in examples:
                print(ex)
        print("=" * 70)

    # =========================================================================
    # plot()
    # =========================================================================

    def plot(self, filename: str = "sp_convergence.png"):
        """
        HSP-style convergence and summary plot.
        Row 0: Bad / Normal / Good path timelines across all 60 months.
        Row 1: Expected Objective | Avg Binary Deviation | Summary.
        """
        n_scenarios   = self.n_scenarios
        is_binary     = self.is_binary
        bin_consensus = self.bin_consensus
        n_binary_na   = self.n_binary_na

        fig, axes = plt.subplots(2, 3, figsize=(21, 10))

        # ------------------------------------------------------------------
        # Shared helpers for the three timeline subplots
        # ------------------------------------------------------------------
        _ref       = self.milp_objects[0]
        _units_all = _ref.U
        _loc_of    = _ref.loc_of
        _loc_names = sorted(set(_loc_of.values()))
        _palette   = ["tab:blue", "tab:orange", "tab:green",
                      "tab:purple", "tab:brown", "tab:pink"]
        _loc_col   = {ln: _palette[i] for i, ln in enumerate(_loc_names)}
        _unit_to_y = {u: i for i, u in enumerate(_units_all)}
        _n_units   = len(_units_all)
        T          = self.T
        _BH        = 0.45
        _s3_start  = self.stage3_months[0] if self.stage3_months else T

        def _short_lbl(u):
            lp, up = u.split(" :: ")
            return f"L{lp.split()[-1]}:U{up.split()[-1]}"

        from matplotlib.lines import Line2D as _L2D
        from matplotlib.patches import Patch as _MPatch

        def _na_decisions(rep_nodes):
            """Extract stocking/harvest consensus decisions for the given NA nodes."""
            stock = {}; harv = {}; hxist = {}
            for _vi, _qn in enumerate(self.all_na_names_ordered):
                if not is_binary[_vi] or bin_consensus.get(_vi, 0.0) < 0.5:
                    continue
                _nk = _qn.split("@")[1]
                if _nk not in rep_nodes:
                    continue
                _vn    = _qn.split("@")[0]
                _br    = _vn.index("[")
                _pfx   = _vn[:_br]
                _inner = _vn[_br + 1:-1]
                if _pfx == "z":
                    _u, _ts = _inner.rsplit(",", 1)
                    stock.setdefault(_u, []).append(int(_ts))
                elif _pfx == "h":
                    _p = _inner.rsplit(",", 2)
                    harv.setdefault(_p[0], []).append((int(_p[1]), int(_p[2])))
                elif _pfx == "h_exist":
                    _u, _ts = _inner.rsplit(",", 1)
                    hxist.setdefault(_u, []).append(int(_ts))
            return stock, harv, hxist

        def _s3_decisions(scenario_name):
            """Extract stage-3 decisions from the post-processing re-solve."""
            s3s = {}; s3h = {}; s3he = {}
            si = next((i for i, nm in enumerate(self.scenario_names)
                       if nm == scenario_name), None)
            if si is None:
                return s3s, s3h, s3he, False
            if self.scenarios[si].SolCount == 0:
                return s3s, s3h, s3he, False
            mlp = self.milp_objects[si]
            for (_u, _ts, _th), _v in mlp.variables["h"].items():
                if _th >= _s3_start:
                    try:
                        if _v.X > 0.5:
                            s3h.setdefault(_u, []).append((_ts, _th))
                    except Exception:
                        pass
            for (_u, _ts), _v in mlp.variables["s"].items():
                if _ts >= _s3_start:
                    try:
                        if _v.X > 0.5:
                            s3s.setdefault(_u, []).append(_ts)
                    except Exception:
                        pass
            for (_u, _th), _v in mlp.variables["h_exist"].items():
                if _th >= _s3_start:
                    try:
                        if _v.X > 0.5:
                            s3he.setdefault(_u, []).append(_th)
                    except Exception:
                        pass
            return s3s, s3h, s3he, True

        def _draw_timeline(ax, rep_nodes, s3_scenario_name, title, shade_color):
            """Draw a complete 60-month stocking/harvest timeline for one path."""
            stock, harv, hxist = _na_decisions(rep_nodes)
            s3s, s3h, s3he, s3_ok = _s3_decisions(s3_scenario_name)

            ax.axvspan(_s3_start, T, alpha=0.07, color=shade_color, zorder=0)
            ax.text((_s3_start + T) / 2, _n_units - 0.15,
                    "stage-3" if s3_ok else "stage-3\n(no sol)",
                    ha="center", va="top", fontsize=6, color=shade_color, style="italic")
            for _sb, _sl in [(self.stage0_end, "S1"),
                              (self.stage1_end, "S2"),
                              (self.stage2_end, "S3")]:
                ax.axvline(_sb, color="gray", ls="--", lw=0.8, alpha=0.6)
                ax.text(_sb + 0.4, _n_units - 0.1, _sl,
                        fontsize=6, color="gray", va="top")

            for _u in _units_all:
                _y   = _unit_to_y[_u]
                _col = _loc_col[_loc_of[_u]]
                # Existing cohort
                if _u in _ref.U_exist:
                    _ehs = hxist.get(_u, [])
                    _s3_ehs = s3he.get(_u, [])
                    if _ehs:
                        for _th in _ehs:
                            ax.barh(_y, _th, left=0, height=_BH,
                                    color=_col, alpha=0.25, edgecolor="none", zorder=1)
                            ax.plot(_th, _y, "v", color="darkorange", ms=6, zorder=4)
                    elif _s3_ehs:
                        ax.barh(_y, _s3_start, left=0, height=_BH,
                                color=_col, alpha=0.10, edgecolor="none", zorder=1)
                    else:
                        # No harvest at all — extend to horizon end
                        ax.barh(_y, T, left=0, height=_BH,
                                color=_col, alpha=0.10, edgecolor="none", zorder=1)
                    for _th in _s3_ehs:
                        ax.barh(_y, _th, left=0, height=_BH,
                                color=_col, alpha=0.25, edgecolor="none", zorder=1)
                        ax.plot(_th, _y, "v", color="darkorange", ms=6, zorder=4)
                # New cohorts — NA stages 0-2
                for (_s0, _th) in harv.get(_u, []):
                    ax.barh(_y, _th - _s0, left=_s0, height=_BH,
                            color=_col, alpha=0.55, edgecolor=_col, lw=0.5, zorder=2)
                    ax.plot(_th, _y, "v", color="tab:red", ms=5, zorder=4)
                for _t in stock.get(_u, []):
                    ax.plot(_t, _y, "^", color="tab:green", ms=6, zorder=5)
                # New cohorts — stage-3 (dashed, slightly transparent)
                for (_s0, _th) in s3h.get(_u, []):
                    ax.barh(_y, _th - _s0, left=_s0, height=_BH,
                            color=_col, alpha=0.35, edgecolor=_col, lw=0.5,
                            linestyle="--", zorder=2)
                    ax.plot(_th, _y, "v", color="tab:red", ms=5, zorder=4, alpha=0.7)
                for _t in s3s.get(_u, []):
                    ax.plot(_t, _y, "^", color="tab:green", ms=6, zorder=5, alpha=0.7)

                # Terminal cohorts: stocked but not harvested within the horizon —
                # draw a dotted bar from stock time to the horizon end.
                _harvested_starts_na  = {_s0 for (_s0, _) in harv.get(_u, [])}
                _harvested_starts_s3  = {_s0 for (_s0, _) in s3h.get(_u, [])}
                _all_harvested_starts = _harvested_starts_na | _harvested_starts_s3
                for _t in stock.get(_u, []):
                    if _t not in _all_harvested_starts:
                        ax.barh(_y, T - _t, left=_t, height=_BH,
                                color=_col, alpha=0.20, edgecolor=_col, lw=0.5,
                                linestyle=":", zorder=1)
                for _t in s3s.get(_u, []):
                    if _t not in _all_harvested_starts:
                        ax.barh(_y, T - _t, left=_t, height=_BH,
                                color=_col, alpha=0.15, edgecolor=_col, lw=0.5,
                                linestyle=":", zorder=1)

            ax.set_xlim(0, T); ax.set_ylim(-0.7, _n_units - 0.3)
            ax.set_yticks(range(_n_units))
            ax.set_yticklabels([_short_lbl(u) for u in _units_all], fontsize=7)
            ax.set_xlabel("Month"); ax.set_title(title)
            ax.grid(True, axis="x", alpha=0.25)
            _leg_h = [
                _L2D([0], [0], marker="^", color="w",
                     markerfacecolor="tab:green",  ms=7, label="Stocking"),
                _L2D([0], [0], marker="v", color="w",
                     markerfacecolor="tab:red",    ms=7, label="Harvest (new)"),
                _L2D([0], [0], marker="v", color="w",
                     markerfacecolor="darkorange", ms=7, label="Harvest (exist.)"),
                _MPatch(color="gray", alpha=0.20, linestyle=":",
                        label="Terminal (unharvested)"),
            ] + [_MPatch(color=_palette[i], alpha=0.55,
                         label=ln.replace("Location ", "Loc"))
                 for i, ln in enumerate(_loc_names)]
            ax.legend(handles=_leg_h, fontsize=6, loc="upper right", ncol=1)

        # ------------------------------------------------------------------
        # Row 0: Bad / Normal / Good path timelines
        # ------------------------------------------------------------------
        _draw_timeline(
            axes[0, 0],
            rep_nodes={"s0", "s1_bad", "s2_bad_bad"},
            s3_scenario_name="s1_bad__s2_bad__s3_bad__s4_bad",
            title="Bad Path  (s0 → s1_bad → s2_bad → s3_bad)",
            shade_color="tab:red",
        )
        _draw_timeline(
            axes[0, 1],
            rep_nodes={"s0", "s1_normal", "s2_normal_normal"},
            s3_scenario_name="s1_normal__s2_normal__s3_normal__s4_normal",
            title="Normal Path  (s0 → s1_normal → s2_normal → s3_normal)",
            shade_color="steelblue",
        )
        _draw_timeline(
            axes[0, 2],
            rep_nodes={"s0", "s1_good", "s2_good_good"},
            s3_scenario_name="s1_good__s2_good__s3_good__s4_good",
            title="Good Path  (s0 → s1_good → s2_good → s3_good)",
            shade_color="tab:green",
        )

        # ------------------------------------------------------------------
        # Row 1: Expected Objective | Avg Binary Deviation | Summary
        # ------------------------------------------------------------------
        ax = axes[1, 0]
        ax.plot(list(range(len(self.objective_history))), self.objective_history,
                "s-", color="tab:blue", markersize=3)
        ax.set_xlabel("Iteration"); ax.set_ylabel("Expected Objective (NOK)")
        ax.set_title("Expected Objective"); ax.grid(True, alpha=0.3)

        ax = axes[1, 1]
        ax.plot(list(range(len(self.avg_dev_bin_history))), self.avg_dev_bin_history,
                "o-", color="tab:orange", markersize=3, label="Avg binary dev")
        ax.axhline(self.epsilon_bin, color="gray", ls="--", alpha=0.6, label="ε_bin")
        ax.set_xlabel("Iteration"); ax.set_ylabel("Avg Abs Deviation")
        ax.set_title("Avg Deviation: Binary NA Vars"); ax.legend(); ax.grid(True, alpha=0.3)

        ax = axes[1, 2]
        feasibility_str = (
            f"{self.n_feasible}/{n_scenarios} feasible"
            + (" [ALL OK]" if self.n_infeasible == 0 else f" [{self.n_infeasible} INFEASIBLE]")
        )
        ax.text(0.5, 0.5,
                f"Total PH time: {self.total_time:.1f}s\n"
                f"Iterations: {self.n_iters}\n"
                f"Avg per iter: {self.total_time / max(self.n_iters, 1):.1f}s\n"
                f"Variables fixed: {self.fixed.sum()}/{n_binary_na}\n"
                f"Final avg bin dev: {self.avg_dev_bin:.4f}\n"
                f"Converged: {self.converged}\n\n"
                f"--- Tree-aware NA ---\n"
                f"s0: 81 scens | s1: 27/node | s2: 9/node\n"
                f"Total NA vars: {self.V} ({n_binary_na} binary)\n\n"
                f"--- Post-processing ---\n"
                f"Feasibility: {feasibility_str}\n"
                f"E[obj] (feasible): {self.eval_obj:,.0f}\n"
                f"E[obj] PH aux:     {self.objective_history[-1]:,.0f}",
                transform=ax.transAxes, fontsize=11, va="center", ha="center",
                bbox=dict(boxstyle="round", facecolor="wheat", alpha=0.5))
        ax.set_title("Summary"); ax.axis("off")

        fig.tight_layout()
        plt.savefig(filename, dpi=150)
        print(f"Plot saved to {filename}")
        plt.close(fig)

        # ------------------------------------------------------------------
        # Biomass trajectory plot (separate figure)
        # ------------------------------------------------------------------
        biomass_filename = filename.replace(".png", "_biomass.png")
        self._plot_biomass(biomass_filename)


    def _plot_biomass(self, filename: str = "sp_biomass.png"):
        """
        Plot total regional biomass trajectories across all 81 scenarios.
        Gray lines for individual scenarios, blue line for expected (prob-weighted) biomass,
        red dashed line for regional MAB limit, purple dashed lines for stage boundaries.
        """
        n_scenarios = self.n_scenarios
        T = self.T
        probs = self.probs

        # Compute total regional biomass per scenario per month
        biomass_data = {s: np.zeros(T) for s in range(n_scenarios)}

        for s in range(n_scenarios):
            milp = self.milp_objects[s]
            model = self.scenarios[s]
            if model.SolCount == 0:
                continue

            for u in milp.U:
                for t in range(T):
                    # New cohorts
                    for ss in milp.A_by_ut.get((u, t), []):
                        if (u, ss, t) in milp.A_harv_set:
                            harvested_through_t = sum(
                                milp.variables["x"][u, ss, tau].X
                                for tau in milp.H_by_us[(u, ss)] if tau <= t
                            )
                            biomass_data[s][t] += float(milp.bpf[ss][t]) * (
                                milp.variables["q"][u, ss].X - harvested_through_t
                            )
                    # Existing cohorts
                    if u in milp.U_exist:
                        try:
                            e_val = milp.variables["e_alive"][u, t].X
                            h_val = milp.variables["h_exist"][u, t].X
                            biomass_data[s][t] += milp.b_exist[u][t] * (e_val - h_val)
                        except Exception:
                            pass

        # Build plot
        months = list(range(T))
        fig, ax = plt.subplots(figsize=(14, 6))

        # Individual scenario lines (gray)
        for s in range(n_scenarios):
            ax.plot(months, biomass_data[s] / 1e3, "-", color="gray", lw=0.8, alpha=0.3)

        # Expected biomass (probability-weighted)
        exp_bio = np.zeros(T)
        for s in range(n_scenarios):
            exp_bio += probs[s] * biomass_data[s]
        ax.plot(months, exp_bio / 1e3, "o-", color="tab:blue", lw=2.5, ms=4,
                zorder=10, label="Expected biomass")

        # Regional MAB limit
        reg_mab_t = self.regional_mab / 1e3
        ax.axhline(reg_mab_t, color="red", ls="--", lw=1.5, alpha=0.7,
                   label=f"Regional MAB ({reg_mab_t:,.0f} t)")

        # Stage boundaries
        for bnd in [self.stage0_end, self.stage1_end, self.stage2_end]:
            ax.axvline(bnd, color="purple", ls="--", lw=0.6, alpha=0.3)

        ax.set_xlabel("Month")
        ax.set_ylabel("Total Biomass (tonnes)")
        ax.set_title("Regional Biomass Trajectories (all 81 scenarios)")
        ax.grid(True, alpha=0.3)
        ax.set_xlim(0, T)
        ax.legend(fontsize=9, loc="upper right")

        fig.tight_layout()
        plt.savefig(filename, dpi=150)
        print(f"Biomass plot saved to {filename}")
        plt.close(fig)

    def plot_cohort_timeline(self, filename: str = "sp_cohort_normal_path.png"):
        """
        Standalone large cohort timeline for the Normal Path representative solution.
        Draws stage-0/1/2 NA decisions for the all-normal node sequence plus stage-3
        decisions from the all-normal scenario.  Suitable for inclusion in a report.
        """
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        from matplotlib.lines import Line2D as _L2D
        from matplotlib.patches import Patch as _MPatch

        is_binary     = self.is_binary
        bin_consensus = self.bin_consensus
        T             = self.T
        _ref          = self.milp_objects[0]
        _units_all    = _ref.U
        _loc_of       = _ref.loc_of
        _loc_names    = sorted(set(_loc_of.values()))
        _palette      = ["tab:blue", "tab:orange", "tab:green",
                         "tab:purple", "tab:brown", "tab:pink"]
        _loc_col      = {ln: _palette[i] for i, ln in enumerate(_loc_names)}
        _unit_to_y    = {u: i for i, u in enumerate(_units_all)}
        _n_units      = len(_units_all)
        _BH           = 0.45
        _s3_start     = self.stage3_months[0] if self.stage3_months else T

        def _short_lbl(u):
            lp, up = u.split(" :: ")
            return f"L{lp.split()[-1]}:U{up.split()[-1]}"

        def _na_decisions(rep_nodes):
            stock = {}; harv = {}; hxist = {}
            for _vi, _qn in enumerate(self.all_na_names_ordered):
                if not is_binary[_vi] or bin_consensus.get(_vi, 0.0) < 0.5:
                    continue
                _nk = _qn.split("@")[1]
                if _nk not in rep_nodes:
                    continue
                _vn    = _qn.split("@")[0]
                _br    = _vn.index("[")
                _pfx   = _vn[:_br]
                _inner = _vn[_br + 1:-1]
                if _pfx == "z":
                    _u, _ts = _inner.rsplit(",", 1)
                    stock.setdefault(_u, []).append(int(_ts))
                elif _pfx == "h":
                    _p = _inner.rsplit(",", 2)
                    harv.setdefault(_p[0], []).append((int(_p[1]), int(_p[2])))
                elif _pfx == "h_exist":
                    _u, _ts = _inner.rsplit(",", 1)
                    hxist.setdefault(_u, []).append(int(_ts))
            return stock, harv, hxist

        def _s3_decisions(scenario_name):
            s3s = {}; s3h = {}; s3he = {}
            si = next((i for i, nm in enumerate(self.scenario_names)
                       if nm == scenario_name), None)
            if si is None:
                return s3s, s3h, s3he, False
            if self.scenarios[si].SolCount == 0:
                return s3s, s3h, s3he, False
            mlp = self.milp_objects[si]
            for (_u, _ts, _th), _v in mlp.variables["h"].items():
                if _th >= _s3_start:
                    try:
                        if _v.X > 0.5:
                            s3h.setdefault(_u, []).append((_ts, _th))
                    except Exception:
                        pass
            for (_u, _ts), _v in mlp.variables["s"].items():
                if _ts >= _s3_start:
                    try:
                        if _v.X > 0.5:
                            s3s.setdefault(_u, []).append(_ts)
                    except Exception:
                        pass
            for (_u, _th), _v in mlp.variables["h_exist"].items():
                if _th >= _s3_start:
                    try:
                        if _v.X > 0.5:
                            s3he.setdefault(_u, []).append(_th)
                    except Exception:
                        pass
            return s3s, s3h, s3he, True

        rep_nodes       = {"s0", "s1_normal", "s2_normal_normal"}
        s3_scenario     = "s1_normal__s2_normal__s3_normal__s4_normal"
        stock, harv, hxist             = _na_decisions(rep_nodes)
        s3s, s3h, s3he, s3_ok         = _s3_decisions(s3_scenario)

        fig, ax = plt.subplots(figsize=(18, max(6, _n_units * 0.55 + 2)))
        shade_color = "steelblue"

        ax.axvspan(_s3_start, T, alpha=0.07, color=shade_color, zorder=0)
        ax.text((_s3_start + T) / 2, _n_units - 0.15,
                "Stage 3" if s3_ok else "Stage 3\n(no sol)",
                ha="center", va="top", fontsize=9, color=shade_color, style="italic")

        stage_labels = [(self.stage0_end, "Stage 1"), (self.stage1_end, "Stage 2"),
                        (self.stage2_end, "Stage 3")]
        for _sb, _sl in stage_labels:
            ax.axvline(_sb, color="gray", ls="--", lw=1.0, alpha=0.7)
            ax.text(_sb + 0.5, _n_units - 0.15, _sl,
                    fontsize=9, color="gray", va="top", fontweight="bold")

        for _u in _units_all:
            _y   = _unit_to_y[_u]
            _col = _loc_col[_loc_of[_u]]
            if _u in _ref.U_exist:
                _ehs   = hxist.get(_u, [])
                _s3_ehs = s3he.get(_u, [])
                if _ehs:
                    for _th in _ehs:
                        ax.barh(_y, _th, left=0, height=_BH,
                                color=_col, alpha=0.25, edgecolor="none", zorder=1)
                        ax.plot(_th, _y, "v", color="darkorange", ms=8, zorder=4)
                elif _s3_ehs:
                    ax.barh(_y, _s3_start, left=0, height=_BH,
                            color=_col, alpha=0.10, edgecolor="none", zorder=1)
                else:
                    ax.barh(_y, T, left=0, height=_BH,
                            color=_col, alpha=0.10, edgecolor="none", zorder=1)
                for _th in _s3_ehs:
                    ax.barh(_y, _th, left=0, height=_BH,
                            color=_col, alpha=0.25, edgecolor="none", zorder=1)
                    ax.plot(_th, _y, "v", color="darkorange", ms=8, zorder=4)
            for (_s0, _th) in harv.get(_u, []):
                ax.barh(_y, _th - _s0, left=_s0, height=_BH,
                        color=_col, alpha=0.55, edgecolor=_col, lw=0.7, zorder=2)
                ax.plot(_th, _y, "v", color="tab:red", ms=7, zorder=4)
            for _t in stock.get(_u, []):
                ax.plot(_t, _y, "^", color="tab:green", ms=8, zorder=5)
            for (_s0, _th) in s3h.get(_u, []):
                ax.barh(_y, _th - _s0, left=_s0, height=_BH,
                        color=_col, alpha=0.35, edgecolor=_col, lw=0.7,
                        linestyle="--", zorder=2)
                ax.plot(_th, _y, "v", color="tab:red", ms=7, zorder=4, alpha=0.7)
            for _t in s3s.get(_u, []):
                ax.plot(_t, _y, "^", color="tab:green", ms=8, zorder=5, alpha=0.7)

            _harvested_na  = {_s0 for (_s0, _) in harv.get(_u, [])}
            _harvested_s3  = {_s0 for (_s0, _) in s3h.get(_u, [])}
            _all_harv      = _harvested_na | _harvested_s3
            for _t in stock.get(_u, []):
                if _t not in _all_harv:
                    ax.barh(_y, T - _t, left=_t, height=_BH,
                            color=_col, alpha=0.20, edgecolor=_col, lw=0.5,
                            linestyle=":", zorder=1)
            for _t in s3s.get(_u, []):
                if _t not in _all_harv:
                    ax.barh(_y, T - _t, left=_t, height=_BH,
                            color=_col, alpha=0.15, edgecolor=_col, lw=0.5,
                            linestyle=":", zorder=1)

        ax.set_xlim(0, T)
        ax.set_ylim(-0.7, _n_units - 0.3)
        ax.set_yticks(range(_n_units))
        ax.set_yticklabels([_short_lbl(u) for u in _units_all], fontsize=10)
        ax.set_xticks(range(0, T + 1, 5))
        ax.set_xlabel("Month", fontsize=12)
        ax.set_title(
            "Representative Production Plan — Normal Temperature Path\n"
            "(s0 → s1_normal → s2_normal → s3_normal; Stage-3 from all-normal scenario)",
            fontsize=13,
        )
        ax.grid(True, axis="x", alpha=0.3)

        _leg_h = [
            _L2D([0], [0], marker="^", color="w",
                 markerfacecolor="tab:green",  ms=9, label="Stocking"),
            _L2D([0], [0], marker="v", color="w",
                 markerfacecolor="tab:red",    ms=9, label="Harvest (new cohort)"),
            _L2D([0], [0], marker="v", color="w",
                 markerfacecolor="darkorange", ms=9, label="Harvest (existing cohort)"),
            _MPatch(color="gray", alpha=0.20, linestyle=":",
                    label="Terminal — stocked, not yet harvested"),
            _MPatch(color="gray", alpha=0.07,
                    label="Stage-3 region (scenario-specific decisions)"),
        ] + [_MPatch(color=_palette[i], alpha=0.55,
                     label=ln.replace("Location ", "Loc "))
             for i, ln in enumerate(_loc_names)]
        ax.legend(handles=_leg_h, fontsize=10, loc="upper right",
                  ncol=2, framealpha=0.9)

        fig.tight_layout()
        plt.savefig(filename, dpi=200)
        print(f"Cohort timeline saved to {filename}")
        plt.close(fig)

    def plot_normal_path_timeline(self, filename: str = "sp_normal_path_timeline.png"):
        """
        Gantt-style cohort timeline for the all-normal scenario path.
        Reads all stocking and harvest decisions directly from the
        s1_normal__s2_normal__s3_normal__s4_normal scenario solution.
        Styled like fh_plan_timeline.png — no special shading for any stage region.
        """
        import matplotlib
        matplotlib.use("Agg")
        import matplotlib.pyplot as plt
        import matplotlib.patches as mpatches
        from matplotlib.lines import Line2D

        normal_name = "s1_normal__s2_normal__s3_normal__s4_normal"
        si = next((i for i, nm in enumerate(self.scenario_names) if nm == normal_name), None)
        if si is None:
            print(f"Scenario '{normal_name}' not found.")
            return
        if self.scenarios[si].SolCount == 0:
            print(f"No solution for scenario '{normal_name}'.")
            return

        mlp = self.milp_objects[si]
        T   = self.T
        T_end = T - 1

        # Extract stocking decisions: z[u, t] stored as variables["s"][(u, t)]
        stocked = set()
        for (u, t), v in mlp.variables["s"].items():
            try:
                if v.X > 0.5:
                    stocked.add((u, t))
            except Exception:
                pass

        # Extract new-cohort harvests: h[u, s_start, t_harv]
        harvests = {}  # (u, s_start) -> t_harv
        for (u, s_start, t_harv), v in mlp.variables["h"].items():
            try:
                if v.X > 0.5:
                    harvests[(u, s_start)] = t_harv
            except Exception:
                pass

        # Extract existing-cohort harvests: h_exist[u, t_harv]
        exist_harvests = {}  # u -> t_harv
        for (u, t_harv), v in mlp.variables["h_exist"].items():
            try:
                if v.X > 0.5:
                    exist_harvests[u] = t_harv
            except Exception:
                pass

        # Unit ordering and location colours
        units_all = mlp.U  # U_exist + U_empty in declaration order
        loc_of    = mlp.loc_of
        loc_names = sorted(set(loc_of.values()))
        palette   = ["tab:blue", "tab:orange", "tab:green",
                     "tab:purple", "tab:brown", "tab:pink"]
        loc_col   = {ln: palette[i] for i, ln in enumerate(loc_names)}

        def _lbl(u):
            lp, up = u.split(" :: ")
            return f"Loc{lp.split()[-1]}-U{up.split()[-1]}"

        n_units   = len(units_all)
        unit_to_y = {u: i for i, u in enumerate(units_all)}
        bar_h     = 0.55
        fig_h     = max(8, n_units * 0.32 + 2)
        fig, ax   = plt.subplots(figsize=(16, fig_h))

        for u in units_all:
            y   = unit_to_y[u]
            col = loc_col[loc_of[u]]

            # Existing cohort bar (starts at 0)
            if u in mlp.U_exist:
                t_h = exist_harvests.get(u)
                bar_end = t_h if t_h is not None else T
                ax.barh(y, bar_end, left=0, height=bar_h,
                        color=col, alpha=0.75, edgecolor="white", linewidth=0.8)
                ax.scatter([0], [y], marker="<", color="steelblue", s=45, zorder=5)
                is_term = t_h is None or t_h >= T_end
                ax.scatter([bar_end], [y], marker="v",
                           color="red" if is_term else "darkorange", s=45, zorder=5)

            # New-cohort bars — one per stocking event on this unit
            for t_s in sorted(t for (uu, t) in stocked if uu == u):
                t_h = harvests.get((u, t_s))
                bar_end = t_h if t_h is not None else T
                ax.barh(y, bar_end - t_s, left=t_s, height=bar_h,
                        color=col, alpha=0.75, edgecolor="white", linewidth=0.8)
                ax.scatter([t_s], [y], marker="^", color="darkgreen", s=45, zorder=5)
                is_term = t_h is None or t_h >= T_end
                ax.scatter([bar_end], [y], marker="v",
                           color="red" if is_term else "darkorange", s=45, zorder=5)

        ax.set_yticks(range(n_units))
        ax.set_yticklabels([_lbl(u) for u in units_all], fontsize=7)
        ax.invert_yaxis()
        ax.set_xlabel("Month", fontsize=11)
        ax.set_xlim(0, T)
        ax.set_xticks(range(0, T + 1, 5))
        ax.grid(True, axis="x", alpha=0.3)

        for bnd in [self.stage0_end, self.stage1_end, self.stage2_end]:
            ax.axvline(bnd, color="gray", ls="--", lw=0.8, alpha=0.5)

        ax.set_title(
            "Stocking & Harvest Timeline — SP Normal Path\n"
            "(months 0–60, all-normal scenario, binary plan)",
            fontsize=12,
        )

        legend_els = [
            mpatches.Patch(color=loc_col[ln], alpha=0.75,
                           label=f"Loc{ln.split()[-1]}")
            for ln in loc_names
        ] + [
            Line2D([0], [0], marker="^", color="w", markerfacecolor="darkgreen",
                   markersize=8, label="Stocking"),
            Line2D([0], [0], marker="<", color="w", markerfacecolor="steelblue",
                   markersize=8, label="Existing/inherited"),
            Line2D([0], [0], marker="v", color="w", markerfacecolor="darkorange",
                   markersize=8, label="Harvest"),
            Line2D([0], [0], marker="v", color="w", markerfacecolor="red",
                   markersize=8, label="Terminal harvest"),
        ]
        ax.legend(handles=legend_els, loc="upper right", fontsize=8, ncol=3, framealpha=0.9)

        fig.tight_layout()
        plt.savefig(filename, dpi=150, bbox_inches="tight")
        plt.close(fig)
        print(f"Normal path timeline saved to {filename}")


def export_solution_bundle(ald: BinaryProgressiveHedging,
                           outdir: str = "solution_exports"):
    """Export final PH policy and scenario solutions to CSV files."""
    os.makedirs(outdir, exist_ok=True)

    # 1) NA policy table (includes @node in var_name for traceability)
    na_rows = []
    for i, qname in enumerate(ald.all_na_names_ordered):
        na_rows.append({
            "na_idx":          i,
            "var_name":        qname.split("@")[0],   # Gurobi variable name
            "node":            qname.split("@")[1],   # node key
            "var_family":      qname.split("[", 1)[0].split("@")[0],
            "consensus_value": float(ald.bin_consensus.get(i, np.nan)),
            "hard_fixed":      int(bool(ald.fixed[i])),
            "fix_value":       float(ald.fix_val[i]),
            "x_bar_value":     float(ald.x_bar[i]) if ald.x_bar is not None else np.nan,
        })
    pd.DataFrame(na_rows).to_csv(os.path.join(outdir, "policy_na.csv"), index=False)

    # 2) Robust q policy
    q_rows = [
        {"node": nk, "unit": u, "stock_month": int(ss), "q_robust": float(v)}
        for (nk, u, ss), v in ald.q_robust_vals.items()
    ]
    pd.DataFrame(q_rows).to_csv(os.path.join(outdir, "robust_q.csv"), index=False)

    # 3) Full scenario variable dump (long format)
    long_rows = []
    for s, milp in enumerate(ald.milp_objects):
        sname   = ald.scenario_names[s]
        has_sol = milp.model.SolCount > 0
        for family, vardict in milp.variables.items():
            for key, var in vardict.items():
                if isinstance(key, tuple):
                    key_pad = list(key) + [None, None, None]
                    k1, k2, k3 = key_pad[:3]
                else:
                    k1, k2, k3 = key, None, None
                long_rows.append({
                    "scenario_id":   s,
                    "scenario_name": sname,
                    "var_family":    family,
                    "k1": k1, "k2": k2, "k3": k3,
                    "value": float(var.X) if has_sol else np.nan,
                })
    pd.DataFrame(long_rows).to_csv(os.path.join(outdir, "scenario_vars_long.csv"), index=False)

    # 4) PH primal matrix (scenario x NA vars) — uses qualified @node names as columns
    ph_sol_df = pd.DataFrame(ald.sol_matrix, columns=ald.all_na_names_ordered)
    ph_sol_df.insert(0, "scenario_id", np.arange(ald.n_scenarios))
    ph_sol_df.to_csv(os.path.join(outdir, "ph_sol_matrix.csv"), index=False)

    # 5) PH multipliers matrix
    ph_lam_df = pd.DataFrame(ald.multipliers, columns=ald.all_na_names_ordered)
    ph_lam_df.insert(0, "scenario_id", np.arange(ald.n_scenarios))
    ph_lam_df.to_csv(os.path.join(outdir, "ph_multipliers.csv"), index=False)

    # 6) Scenario-level objective/feasibility summary
    obj_rows = []
    for s, model in enumerate(ald.scenarios):
        feasible = model.SolCount > 0
        obj_rows.append({
            "scenario_id":   s,
            "scenario_name": ald.scenario_names[s],
            "feasible":      int(feasible),
            "obj_value":     float(model.ObjVal) if feasible else np.nan,
        })
    pd.DataFrame(obj_rows).to_csv(os.path.join(outdir, "scenario_objectives.csv"), index=False)

    print(f"Exported solution bundle to {outdir}")


# =============================================================================
# export_decision_tree — Excel workbook 
# =============================================================================

def export_decision_tree(ald: "BinaryProgressiveHedging",
                         filename: str = "st_sp_decision_tree_full.xlsx"):
    """
    Export the ALD (ST_SP) decision tree to a readable Excel workbook.

    Sheets
    ------
    Tree_Plan        — 40 node-sections (1 stage-0 + 3 stage-1 + 9 stage-2 + 27 stage-3),
                       each listing only the active decisions with quantities.
    Scenario_Lookup  — 81 rows: which nodes apply + feasibility + obj value.
    Scenario_Playbook— Full action list per scenario (all 4 stages).
    Guide            — How to read the workbook.
    """
    try:
        import openpyxl
        from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
        from openpyxl.utils import get_column_letter
        _have_openpyxl = True
    except ImportError:
        _have_openpyxl = False

    labels = ald.labels  # ["bad", "normal", "good"]

    # ------------------------------------------------------------------
    # Build consensus lookup: node_key → {vname: 0/1}
    # ------------------------------------------------------------------
    consensus_by_node: dict[str, dict[str, int]] = {}
    for vi, qname in enumerate(ald.all_na_names_ordered):
        if not ald.is_binary[vi]:
            continue
        nk    = qname.split("@")[1]
        vname = qname.split("@")[0]
        val   = int(round(ald.bin_consensus.get(vi, 0.0)))
        consensus_by_node.setdefault(nk, {})[vname] = val

    # ------------------------------------------------------------------
    # Parse scenario name → (t1, t2, t3, t4)
    # ------------------------------------------------------------------
    def _parse_sname(name):
        p = name.split("__")
        return (p[0].split("_", 1)[1], p[1].split("_", 1)[1],
                p[2].split("_", 1)[1], p[3].split("_", 1)[1])

    def _count(nk, pfx):
        return sum(1 for vn, v in consensus_by_node.get(nk, {}).items()
                   if v == 1 and vn.split("[")[0] == pfx)

    # ------------------------------------------------------------------
    # Scenario_Lookup
    # ------------------------------------------------------------------
    lookup_rows = []
    for s, name in enumerate(ald.scenario_names):
        t1, t2, t3, t4 = _parse_sname(name)
        s1_node = f"s1_{t1}"
        s2_node = f"s2_{t1}_{t2}"
        s3_node = f"s3_{t1}_{t2}_{t3}"
        milp    = ald.milp_objects[s]
        obj_val = milp.model.ObjVal if milp.model.SolCount > 0 else None
        lookup_rows.append({
            "scenario_id":   s,
            "t1 (stage0)":   t1,
            "t2 (stage1)":   t2,
            "t3 (stage2)":   t3,
            "t4 (stage3)":   t4,
            "stage1_node":   s1_node,
            "stage2_node":   s2_node,
            "stage3_node":   s3_node,
            "s0_stocks":     _count("s0",    "z"),
            "s0_harvests":   _count("s0",    "h") + _count("s0", "h_exist"),
            "s1_stocks":     _count(s1_node, "z"),
            "s1_harvests":   _count(s1_node, "h"),
            "s2_stocks":     _count(s2_node, "z"),
            "s2_harvests":   _count(s2_node, "h"),
            "s3_stocks":     _count(s3_node, "z"),
            "s3_harvests":   _count(s3_node, "h"),
            "feasible":      "YES" if s not in ald.infeasible_scenarios else "NO",
            "obj_value":     round(obj_val, 0) if obj_val is not None else "",
        })
    df_lookup = pd.DataFrame(lookup_rows)

    # ------------------------------------------------------------------
    # Scenario_Playbook
    # ------------------------------------------------------------------
    def _stage_from_month(m: int) -> str:
        if m < ald.stage0_end:  return "Stage0"
        if m < ald.stage1_end:  return "Stage1"
        if m < ald.stage2_end:  return "Stage2"
        return "Stage3"

    _stage_order = {"Stage0": 0, "Stage1": 1, "Stage2": 2, "Stage3": 3}

    def _node_key_playbook(stage, t1, t2, sname):
        if stage == "Stage0": return "s0"
        if stage == "Stage1": return f"s1_{t1}"
        if stage == "Stage2": return f"s2_{t1}_{t2}"
        return sname

    playbook_rows = []
    for s, milp in enumerate(ald.milp_objects):
        sname = ald.scenario_names[s]
        t1, t2, t3, t4 = _parse_sname(sname)
        if milp.model.SolCount == 0:
            continue
        q_vars  = milp.variables.get("q", {})
        z_vars  = milp.variables.get("s", {})
        h_vars  = milp.variables.get("h", {})
        hx_vars = milp.variables.get("h_exist", {})
        for (u, ss), zvar in z_vars.items():
            if float(zvar.X) < 0.5:
                continue
            stage = _stage_from_month(int(ss))
            playbook_rows.append({
                "scenario_id": s, "t1": t1, "t2": t2, "t3": t3, "t4": t4,
                "stage": stage, "stage_order": _stage_order[stage],
                "node": _node_key_playbook(stage, t1, t2, sname),
                "action": "Stock", "unit": u,
                "action_month": int(ss), "cohort_start": "",
                "qty_fish": int(round(q_vars[u, ss].X)) if (u, ss) in q_vars else "",
                "weight_per_fish_kg": "",
            })
        for (u, ss, th), hvar in h_vars.items():
            if float(hvar.X) < 0.5:
                continue
            stage = _stage_from_month(int(th))
            _wt_kg = round(milp.wpf[int(ss)][int(th)] / 1000.0, 3)
            playbook_rows.append({
                "scenario_id": s, "t1": t1, "t2": t2, "t3": t3, "t4": t4,
                "stage": stage, "stage_order": _stage_order[stage],
                "node": _node_key_playbook(stage, t1, t2, sname),
                "action": "Harvest", "unit": u,
                "action_month": int(th), "cohort_start": int(ss),
                "qty_fish": "",
                "weight_per_fish_kg": _wt_kg,
            })
        for (u, th), hvar in hx_vars.items():
            if float(hvar.X) < 0.5:
                continue
            stage = _stage_from_month(int(th))
            _wt_kg_ex = round(float(milp.w_exist[u][int(th)]) / 1000.0, 3)
            playbook_rows.append({
                "scenario_id": s, "t1": t1, "t2": t2, "t3": t3, "t4": t4,
                "stage": stage, "stage_order": _stage_order[stage],
                "node": _node_key_playbook(stage, t1, t2, sname),
                "action": "Harvest (existing)", "unit": u,
                "action_month": int(th), "cohort_start": "existing",
                "qty_fish": "",
                "weight_per_fish_kg": _wt_kg_ex,
            })

    df_playbook = (pd.DataFrame(playbook_rows)
                   .sort_values(["scenario_id", "stage_order", "action_month", "unit"])
                   .reset_index(drop=True)
                   if playbook_rows else pd.DataFrame())

    # ------------------------------------------------------------------
    # Guide rows
    # ------------------------------------------------------------------
    guide_rows = [
        {"Sheet": "Tree_Plan",
         "What it shows": "The complete 40-node decision tree.  "
                          "Stage 0 (1 node) → Stage 1 (3 nodes) → Stage 2 (9 nodes) → Stage 3 (27 nodes). "
                          "Each node section lists only the active decisions (Stock / Harvest). "
                          "Stocking quantities (fish) are shown inline in the Qty_Fish column.",
         "How to use": "Read Stage 0 first — those decisions are committed now. "
                       "After month 14, observe t1 and jump to your Stage-1 branch. "
                       "After month 29, observe t2 and jump to your Stage-2 branch. "
                       "After month 44, observe t3 and jump to your Stage-3 branch."},
        {"Sheet": "Scenario_Lookup",
         "What it shows": "81 rows — one per scenario.  Shows which node applies at each stage, "
                          "decision counts per stage, and post-processing feasibility.",
         "How to use": "Find your scenario by t1/t2/t3/t4 to identify which node sections "
                       "of Tree_Plan are relevant to you."},
        {"Sheet": "Scenario_Playbook",
         "What it shows": "Full chronological action list for every scenario across all 4 stages. "
                          "Stage-3 decisions (months 45–59) are scenario-specific and only appear here.",
         "How to use": "Filter by scenario_id to see the exact sequence of Stock / Harvest actions "
                       "for a given scenario, including stage-3 recourse decisions."},
    ]
    df_guide = pd.DataFrame(guide_rows)

    # ------------------------------------------------------------------
    # Write Tree_Plan sheet (section-per-node layout)
    # ------------------------------------------------------------------
    if not _have_openpyxl:
        with pd.ExcelWriter(filename, engine="openpyxl") as writer:
            df_guide.to_excel(writer, sheet_name="Guide", index=False)
            df_lookup.to_excel(writer, sheet_name="Scenario_Lookup", index=False)
            df_playbook.to_excel(writer, sheet_name="Scenario_Playbook", index=False)
        print(f"Decision tree exported to {filename} (no rich formatting — install openpyxl)")
        return

    # ---- Colour palette ------------------------------------------------
    C_STAGE   = PatternFill("solid", fgColor="1F4E79")
    C_NODE    = PatternFill("solid", fgColor="2E75B6")
    C_COL_HDR = PatternFill("solid", fgColor="D6E4F0")
    C_STOCK   = PatternFill("solid", fgColor="E2EFDA")
    C_HARV    = PatternFill("solid", fgColor="FCE4D6")
    C_NONE    = PatternFill("solid", fgColor="F2F2F2")
    C_BLUE_H  = PatternFill("solid", fgColor="BDD7EE")
    C_GREY_H  = PatternFill("solid", fgColor="D9D9D9")

    F_WHITE_BOLD  = Font(bold=True, color="FFFFFF", size=11)
    F_WHITE       = Font(color="FFFFFF", italic=True)
    F_BOLD        = Font(bold=True)
    F_NORMAL      = Font()

    THIN  = Side(style="thin", color="AAAAAA")
    BORD  = Border(left=THIN, right=THIN, bottom=THIN, top=THIN)

    TP_COLS = ["Action", "Unit", "Month", "Cohort_Start", "Qty_Fish", "Harvest_Weight_kg", "Notes"]
    NCOLS   = len(TP_COLS)

    wb    = openpyxl.Workbook()
    ws_tp = wb.active
    ws_tp.title = "Tree_Plan"
    ws_tp.freeze_panes = "A3"

    col_widths = [22, 20, 9, 14, 14, 18, 50]
    for ci, w in enumerate(col_widths, 1):
        ws_tp.column_dimensions[get_column_letter(ci)].width = w

    row = 1

    def _fill_row(ws, r, fill, font=None, values=None):
        for c in range(1, NCOLS + 1):
            cell = ws.cell(row=r, column=c)
            cell.fill = fill
            if font:
                cell.font = font
            cell.border = BORD
            if values and (c - 1) < len(values):
                cell.value = values[c - 1]

    def _write_stage_banner(ws, r, title, subtitle):
        _fill_row(ws, r, C_STAGE, F_WHITE_BOLD)
        ws.cell(row=r, column=1).value = title
        ws.cell(row=r, column=2).value = subtitle
        ws.cell(row=r, column=2).font  = F_WHITE
        ws.row_dimensions[r].height = 18

    def _write_node_banner(ws, r, title, subtitle):
        _fill_row(ws, r, C_NODE, F_WHITE_BOLD)
        ws.cell(row=r, column=1).value = "  " + title
        ws.cell(row=r, column=2).value = subtitle
        ws.cell(row=r, column=2).font  = Font(color="FFFFFF", italic=True)
        ws.row_dimensions[r].height = 16

    def _write_col_headers(ws, r):
        _fill_row(ws, r, C_COL_HDR, F_BOLD, TP_COLS)

    def _write_action_row(ws, r, action, unit, month, cohort_start, qty, harvest_weight, notes):
        fill = C_STOCK if action == "Stock" else C_HARV
        _fill_row(ws, r, fill, F_NORMAL,
                  [action, unit, month, cohort_start,
                   int(qty) if isinstance(qty, float) and not np.isnan(qty) else (qty or ""),
                   harvest_weight,
                   notes])

    def _write_no_decisions(ws, r):
        _fill_row(ws, r, C_NONE)
        ws.cell(row=r, column=1).value = "  (no decisions in this node)"
        ws.cell(row=r, column=1).font  = Font(italic=True, color="888888")

    # Representative milp per node (first scenario in each node) for weight lookups
    _rep_milp_for_node: dict = {}
    for _nk, _nk_scens in ald._node_scens.items():
        if _nk_scens:
            _rep_milp_for_node[_nk] = ald.milp_objects[_nk_scens[0]]

    def _active_actions(nk: str) -> list[dict]:
        acts = []
        _rep_m = _rep_milp_for_node.get(nk)
        for vname, val in sorted(consensus_by_node.get(nk, {}).items()):
            if val != 1:
                continue
            pfx   = vname.split("[")[0]
            inner = vname[vname.index("[")+1:-1]
            if pfx == "z":
                u, ts = inner.rsplit(",", 1)
                ts  = int(ts)
                q   = ald.q_robust_vals.get((nk, u, ts))
                qty_str = f"{int(round(q)):,}" if q is not None else "—"
                acts.append({
                    "action": "Stock", "unit": u, "month": ts,
                    "cohort_start": "", "qty_fish": qty_str,
                    "harvest_weight_kg": "",
                    "notes": f"Stock {u} in month {ts}"
                             + (f"  ({int(round(q)):,} fish)" if q else ""),
                })
            elif pfx == "h":
                p = inner.rsplit(",", 2)
                u, ss, th = p[0], int(p[1]), int(p[2])
                if _rep_m is not None:
                    _wt = round(_rep_m.wpf[ss][th] / 1000.0, 3)
                    _wt_str = f"{_wt:.3f}"
                else:
                    _wt_str = ""
                acts.append({
                    "action": "Harvest", "unit": u, "month": th,
                    "cohort_start": ss, "qty_fish": "",
                    "harvest_weight_kg": _wt_str,
                    "notes": f"Harvest {u} in month {th}  (cohort stocked month {ss})",
                })
            elif pfx == "h_exist":
                u, th = inner.rsplit(",", 1)
                if _rep_m is not None and u in _rep_m.w_exist:
                    _wt_ex = round(float(_rep_m.w_exist[u][int(th)]) / 1000.0, 3)
                    _wt_ex_str = f"{_wt_ex:.3f}"
                else:
                    _wt_ex_str = ""
                acts.append({
                    "action": "Harvest (existing)", "unit": u, "month": int(th),
                    "cohort_start": "existing", "qty_fish": "",
                    "harvest_weight_kg": _wt_ex_str,
                    "notes": f"Harvest existing stock in {u} in month {int(th)}",
                })
        acts.sort(key=lambda a: (0 if a["action"] == "Stock" else 1,
                                 a["month"], a["unit"]))
        return acts

    # ---- STAGE 0 -------------------------------------------------------
    n_s0 = len(ald._node_scens.get("s0", []))
    _write_stage_banner(ws_tp, row,
                        "STAGE 0  —  Months 0–14",
                        f"Committed before any uncertainty resolves.  "
                        f"Identical for ALL {n_s0} scenarios.")
    row += 1
    _write_col_headers(ws_tp, row); row += 1

    for act in _active_actions("s0"):
        _write_action_row(ws_tp, row, act["action"], act["unit"],
                          act["month"], act["cohort_start"],
                          act["qty_fish"], act["harvest_weight_kg"], act["notes"])
        row += 1
    if not _active_actions("s0"):
        _write_no_decisions(ws_tp, row); row += 1
    row += 1

    # ---- STAGE 1 -------------------------------------------------------
    _write_stage_banner(ws_tp, row,
                        "STAGE 1  —  Months 15–29",
                        "Decided after observing temperature outcome in stage 0 (t1).  "
                        "Three branches — follow only the one that matches your observed t1.")
    row += 1

    for t1 in labels:
        nk      = f"s1_{t1}"
        n_scens = len(ald._node_scens.get(nk, []))
        _write_node_banner(ws_tp, row,
                           f"IF  t1 = {t1.upper()}",
                           f"{n_scens} of 81 scenarios follow this branch  "
                           f"→ stage-2 node will be s2_{t1}_{{t2}}")
        row += 1
        _write_col_headers(ws_tp, row); row += 1

        actions = _active_actions(nk)
        for act in actions:
            _write_action_row(ws_tp, row, act["action"], act["unit"],
                              act["month"], act["cohort_start"],
                              act["qty_fish"], act["harvest_weight_kg"], act["notes"])
            row += 1
        if not actions:
            _write_no_decisions(ws_tp, row); row += 1
        row += 1

    # ---- STAGE 2 -------------------------------------------------------
    _write_stage_banner(ws_tp, row,
                        "STAGE 2  —  Months 30–44",
                        "Decided after observing temperature in stage 1 (t2).  "
                        "Nine branches — follow the one matching your (t1, t2) pair.")
    row += 1

    for t1 in labels:
        for t2 in labels:
            nk      = f"s2_{t1}_{t2}"
            n_scens = len(ald._node_scens.get(nk, []))
            _write_node_banner(ws_tp, row,
                               f"IF  t1 = {t1.upper()}  AND  t2 = {t2.upper()}",
                               f"{n_scens} of 81 scenarios follow this branch")
            row += 1
            _write_col_headers(ws_tp, row); row += 1

            actions = _active_actions(nk)
            for act in actions:
                _write_action_row(ws_tp, row, act["action"], act["unit"],
                                  act["month"], act["cohort_start"],
                                  act["qty_fish"], act["harvest_weight_kg"], act["notes"])
                row += 1
            if not actions:
                _write_no_decisions(ws_tp, row); row += 1
            row += 1

    # ---- STAGE 3 (27 nodes — one per (t1, t2, t3) triple) ---------------
    _write_stage_banner(ws_tp, row,
                        "STAGE 3  —  Months 45–59",
                        "Decided after observing t3.  "
                        "27 branches — follow the one matching your (t1, t2, t3) triple.")
    row += 1

    for t1 in labels:
        for t2 in labels:
            for t3 in labels:
                nk      = f"s3_{t1}_{t2}_{t3}"
                n_scens = len(ald._node_scens.get(nk, []))
                _write_node_banner(ws_tp, row,
                                   f"IF  t1={t1.upper()}  t2={t2.upper()}  t3={t3.upper()}",
                                   f"{n_scens} of 81 scenarios follow this branch")
                row += 1
                _write_col_headers(ws_tp, row); row += 1

                actions = _active_actions(nk)
                for act in actions:
                    _write_action_row(ws_tp, row, act["action"], act["unit"],
                                      act["month"], act["cohort_start"],
                                      act["qty_fish"], act["harvest_weight_kg"], act["notes"])
                    row += 1
                if not actions:
                    _write_no_decisions(ws_tp, row); row += 1
                row += 1

    # ------------------------------------------------------------------
    # Append Scenario_Lookup, Scenario_Playbook, Guide sheets
    # ------------------------------------------------------------------
    def _add_df_sheet(wb, df, sheet_name, header_fill):
        ws = wb.create_sheet(title=sheet_name)
        for ci, col in enumerate(df.columns, 1):
            cell = ws.cell(row=1, column=ci, value=col)
            cell.fill = header_fill
            cell.font = F_BOLD
            cell.border = BORD
        for ri, row_data in enumerate(df.itertuples(index=False), start=2):
            for ci, val in enumerate(row_data, start=1):
                cell = ws.cell(row=ri, column=ci,
                               value=(None if (isinstance(val, float) and np.isnan(val)) else val))
                cell.border = BORD
        for col in ws.columns:
            max_len = max(
                (len(str(cell.value)) if cell.value is not None else 0) for cell in col
            )
            ws.column_dimensions[get_column_letter(col[0].column)].width = min(max_len + 3, 45)
        ws.freeze_panes = "A2"
        return ws

    _add_df_sheet(wb, df_lookup,   "Scenario_Lookup",   C_BLUE_H)
    _add_df_sheet(wb, df_playbook, "Scenario_Playbook", C_BLUE_H)

    ws_g = wb.create_sheet(title="Guide")
    for ci, col in enumerate(df_guide.columns, 1):
        cell = ws_g.cell(row=1, column=ci, value=col)
        cell.fill = C_GREY_H; cell.font = F_BOLD; cell.border = BORD
    for ri, row_data in enumerate(df_guide.itertuples(index=False), start=2):
        for ci, val in enumerate(row_data, start=1):
            cell = ws_g.cell(row=ri, column=ci, value=val)
            cell.alignment = Alignment(wrap_text=True, vertical="top")
            cell.border = BORD
    ws_g.column_dimensions["A"].width = 20
    ws_g.column_dimensions["B"].width = 60
    ws_g.column_dimensions["C"].width = 60
    for r in range(2, len(df_guide) + 2):
        ws_g.row_dimensions[r].height = 55

    wb.save(filename)
    print(f"Decision tree exported to {filename}")
    print(f"  Sheets: Tree_Plan | Scenario_Lookup | Scenario_Playbook | Guide")


# =============================================================================
# __main__
# =============================================================================

if __name__ == "__main__":
    from instance import temps_normal_12, temps_bad_12, temps_good_12
    ald = BinaryProgressiveHedging(
        units_df=units_df,
        loc_mab=loc_mab,
        regional_mab=regional_mab,
        temps_bad=temps_bad_12,
        temps_normal=temps_normal_12,
        temps_good=temps_good_12,
        mip_gap=0.02,
        tail_fix_window=5,
        tail_fix_max_disagree=30,
    )
    ald.build()
    ald.solve()
    export_solution_bundle(ald, outdir="solution_exports_st_full_sp")
    export_decision_tree(ald, filename="sp_decision_tree_full.xlsx")
    ald.print_results()
    ald.plot("sp_convergence_full.png")
    ald.plot_normal_path_timeline("sp_normal_path_timeline.png")

# (file touched 2026-05-01 to force OneDrive re-sync of bash mount)
